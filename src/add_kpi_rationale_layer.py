from __future__ import annotations

import csv
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
REPO = OUT / "leatt-fabric-bi-ml-git-proof"
SQLITE = OUT / "leatt_ecommerce_warehouse.sqlite"
MASTER = OUT / "leatt_ecommerce_bi_ml_ai_command_center_master.xlsx"


def money(value: float) -> str:
    return f"R{value:,.0f}"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_values() -> dict[str, str]:
    conn = sqlite3.connect(SQLITE)
    total_revenue, total_margin, total_returns, orders, lines = conn.execute(
        """
        select
            sum(net_revenue_zar),
            sum(gross_margin_zar),
            sum(return_amount_zar),
            count(distinct order_id),
            count(*)
        from fact_transaction_lines
        """
    ).fetchone()
    top_category = conn.execute(
        "select category, net_revenue_zar from agg_category order by net_revenue_zar desc limit 1"
    ).fetchone()
    top_channel = pd.read_csv(OUT / "leatt_intelligent_channel.csv").sort_values("roas", ascending=False).iloc[0]
    ml = read_csv(OUT / "leatt_ml_return_risk_metrics.csv")[0]
    conn.close()
    return {
        "total_revenue": money(total_revenue),
        "gross_margin": money(total_margin),
        "gross_margin_rate": f"{total_margin / total_revenue:.1%}",
        "return_rate": f"{total_returns / total_revenue:.1%}",
        "orders": f"{orders:,}",
        "lines": f"{lines:,}",
        "aov": money(total_revenue / orders),
        "top_category": f"{top_category[0]} ({money(top_category[1])})",
        "top_channel": f"{top_channel['channel']} ({top_channel['roas']:.1f}x ROAS)",
        "ml_auc": f"{float(ml['auc']):.3f}",
        "ml_recall": f"{float(ml['recall']):.3f}",
    }


def build_kpis(values: dict[str, str]) -> pd.DataFrame:
    rows = [
        ("Executive", "Net Revenue", values["total_revenue"], "Sum of transaction net revenue after discounts and excluding returns.", "Top-line demand is the first board-level signal of commercial scale.", "Growth or contraction in market demand, product-market fit and campaign reach.", "Investigate by month, category, channel and province; protect supply for demand peaks.", "Executive Overview, Monthly Trend, AI Command Center"),
        ("Executive", "Gross Margin", values["gross_margin"], "Revenue less modeled product cost and fulfillment economics.", "Revenue without margin can destroy value; margin shows the quality of growth.", "Whether the business is buying growth too expensively or selling profitable baskets.", "Prioritize margin-positive bundles, pricing reviews and channel mix changes.", "Executive Overview, Category KPI, Product KPI"),
        ("Executive", "Gross Margin Rate", values["gross_margin_rate"], "Gross margin divided by net revenue.", "Normalizes margin so categories and channels can be compared fairly.", "Margin compression, discount overuse, poor channel economics or weak product mix.", "Tighten discount rules, promote higher-margin products, renegotiate COGS assumptions.", "Executive Overview, Category KPI, Finance Pack"),
        ("Customer", "Average Order Value", values["aov"], "Net revenue divided by distinct orders.", "AOV shows basket depth and upsell effectiveness.", "Whether customers are buying single items or building richer baskets.", "Launch bundles, recommendations, threshold offers and accessory attach campaigns.", "Marketing Report, A/B Testing, AI Command Center"),
        ("Operations", "Return Rate", values["return_rate"], "Return amount divided by net revenue.", "Returns are direct margin leakage and a customer-experience signal.", "Fit/sizing issues, unclear PDP content, damaged delivery, buyer remorse or quality issues.", "Improve size guides, PDP detail, reviews, post-purchase education and exception reporting.", "ML Monitoring, Accounting Governance, COO Watchlist"),
        ("Merchandising", "Hero Category Revenue", values["top_category"], "Highest revenue category in the modeled catalog and transaction estate.", "Shows which category deserves executive attention and protected availability.", "Where customer demand is concentrated and where stockouts would hurt most.", "Protect inventory, create SEO hubs, develop bundles and track category share monthly.", "Category KPI, SEO Roadmap, Executive Dossier"),
        ("Marketing", "Channel ROAS", values["top_channel"], "Revenue divided by estimated channel spend.", "ROAS controls media efficiency and prevents scaling unprofitable traffic.", "Which acquisition journeys are efficient versus experimental or wasteful.", "Scale high-ROAS journeys; constrain low-ROAS spend until creative and landing pages improve.", "Marketing ROI, Channel Intelligence, A/B Test Pack"),
        ("Marketing", "A/B Test Uplift", "Included in experiment sheets", "Difference in conversion, AOV or margin between control and variant.", "Testing separates opinion from evidence before using Azure credit or media budget.", "Whether a bundle, threshold, copy or landing-page change creates measurable lift.", "Roll out significant winners; keep inconclusive tests as learning only.", "A/B Test Results, Growth Marketing Report"),
        ("ML", "Return-Risk AUC", values["ml_auc"], "Area under ROC curve for the return-risk classifier.", "AUC measures ranking quality across thresholds.", "Whether the model can separate risky orders from normal orders better than chance.", "Use as a watchlist and improve with real return reasons, sizing and delivery signals.", "ML Monitoring, AI Command Center"),
        ("ML", "Return-Risk Recall", values["ml_recall"], "Share of actual return-risk cases captured by the model.", "For operations, missing risky returns is usually worse than reviewing extra orders.", "How much return exposure can be surfaced early.", "Route high-risk cohorts to PDP fixes, customer education and service follow-up.", "ML Monitoring, Product Watchlist"),
        ("Finance", "VAT Reconciliation", "Included in finance pack", "Compares ecommerce sales/refunds/VAT to SAP or accounting postings.", "Finance needs provable close controls, not only dashboard totals.", "Unposted refunds, settlement timing gaps, VAT exceptions or account mapping issues.", "Create exception owners and include reconciliations in monthly close evidence.", "Accounting Governance Pack, SAP Mapping"),
        ("Governance", "Row Count Control", values["lines"], "Source and target row counts for uploaded files, tables and semantic model inputs.", "A simple but powerful audit control that catches missing or duplicated loads.", "Pipeline completeness, accidental truncation, duplicate loads or stale extracts.", "Block refresh promotion when row counts deviate outside tolerance.", "Fabric Proof, Source Register, Governance Scorecard"),
        ("Governance", "Source Lineage", "Documented", "Maps website, generated data, Fabric paths, Power BI model, reports and Git commits.", "Lineage makes the work auditable and easier to operate.", "Whether a KPI can be traced back to a reliable source and transformation route.", "Publish lineage with every report refresh and monthly close pack.", "Source Register, Git Proof, Completion Report"),
        ("Platform", "Fabric Capacity State", "Paused after handover", "Azure Fabric F2 capacity running or paused state.", "Cloud cost control is part of senior BI ownership.", "Whether proof infrastructure is still consuming Azure credit.", "Resume only for refresh/demo; suspend after review.", "Cost Handover, Final Completion Report"),
    ]
    return pd.DataFrame(
        rows,
        columns=["domain", "kpi", "current_value_or_status", "definition", "why_used", "what_it_signals", "business_action", "report_placement"],
    )


def build_reports() -> pd.DataFrame:
    rows = [
        ("Executive Overview", "CEO, CFO, Head of Ecommerce", "Revenue, margin, AOV, return rate, monthly trend, hero categories", "Gives the quickest health check of scale, profitability and trading momentum.", "Use first in project reviews and executive readouts."),
        ("AI Commerce Command Center", "Executive team, BI Ops, Marketing, Merchandising", "Next-best actions, agents, anomalies, category roles, channel efficiency", "Shows the shift from passive reporting to intelligent operating rhythm.", "Use to explain how AI agents can turn BI signals into actions."),
        ("ML Return-Risk Monitoring", "COO, Customer Experience, BI/ML team", "AUC, recall, risk bands, risky categories/products", "Demonstrates applied ML and shows where return leakage can be attacked.", "Use to position model as a watchlist, not a perfect oracle."),
        ("Marketing Growth Report", "CMO, Performance Marketing", "ROAS, A/B uplift, SEO roadmap, competitor comparison, campaign calendar", "Connects BI to revenue growth and Azure-credit discipline.", "Use to justify scaling only after tests prove lift."),
        ("Accounting Governance Pack", "CFO, Finance Manager, Auditor", "VAT, refunds, journal entries, SAP mapping, exceptions", "Shows accounting-grade reconciliation, not just analytics storytelling.", "Use to prove finance, SAP BW/SAP B1 and audit readiness."),
        ("ERD And Fabric Deployment Report", "Data Architect, BI Lead", "Star schema, table grain, relationships, OneLake paths, pipeline proof", "Explains how the data model is built and why it can scale.", "Use during architecture and deployment reviews."),
        ("Power BI Screenshot Pack", "Project stakeholders", "Executive, ML and marketing proof screenshots", "Provides visual evidence even if the live tenant is not open.", "Use as a lightweight project evidence pack."),
        ("Final Executive Dossier", "All stakeholders", "Concise end-to-end story, screenshots, KPIs, handover and cost control", "Consolidates the whole project into one polished artifact.", "Use as the main project summary."),
    ]
    return pd.DataFrame(rows, columns=["report", "audience", "primary_kpis", "why_this_report_exists", "how_to_use_it"])


def markdown_table(df: pd.DataFrame) -> str:
    rows = ["| " + " | ".join(df.columns) + " |", "| " + " | ".join(["---"] * len(df.columns)) + " |"]
    for _, row in df.iterrows():
        rows.append("| " + " | ".join(str(row[col]).replace("\n", " ") for col in df.columns) + " |")
    return "\n".join(rows)


def write_outputs(kpis: pd.DataFrame, reports: pd.DataFrame) -> list[Path]:
    paths: list[Path] = []
    kpi_csv = OUT / "leatt_kpi_rationale_catalog.csv"
    reports_csv = OUT / "leatt_report_rationale_catalog.csv"
    kpis.to_csv(kpi_csv, index=False)
    reports.to_csv(reports_csv, index=False)
    paths.extend([kpi_csv, reports_csv])

    md = OUT / "kpi_and_report_rationale.md"
    md.write_text(
        f"""# KPI And Report Rationale

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}

This section explains the KPIs and reports used in the Leatt Fabric BI/ML portfolio. The goal is to show not only what was measured, but why it matters, what it signals, and which business action it should trigger.

## KPI Catalog

{markdown_table(kpis)}

## Report Catalog

{markdown_table(reports)}

## Design Principle

The project avoids vanity metrics. Every KPI must answer one of four business questions:

- Are we growing?
- Are we growing profitably?
- Where is leakage or risk?
- What action should the business take next?
""",
        encoding="utf-8",
    )
    paths.append(md)

    pdf = OUT / "leatt_kpi_and_report_rationale.pdf"
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleRed", parent=styles["Title"], alignment=TA_CENTER, textColor=colors.HexColor("#d71920")))
    story = [
        Paragraph("KPI And Report Rationale", styles["TitleRed"]),
        Paragraph("Why each KPI was used, what it signals, and what action it drives.", styles["BodyText"]),
        Spacer(1, 0.4 * cm),
        Table(
            [list(kpis.columns)] + kpis.values.tolist(),
            colWidths=[2.3 * cm, 3.2 * cm, 3.0 * cm, 4.0 * cm, 4.2 * cm, 4.2 * cm, 4.2 * cm, 3.2 * cm],
            repeatRows=1,
            style=[
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
                ("FONTSIZE", (0, 0), (-1, -1), 5.6),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ],
        ),
        PageBreak(),
        Paragraph("Report Catalog", styles["Heading1"]),
        Table(
            [list(reports.columns)] + reports.values.tolist(),
            colWidths=[4.0 * cm, 4.0 * cm, 6.0 * cm, 7.0 * cm, 6.0 * cm],
            repeatRows=1,
            style=[
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ],
        ),
    ]
    SimpleDocTemplate(str(pdf), pagesize=landscape(A4), rightMargin=0.7 * cm, leftMargin=0.7 * cm, topMargin=0.7 * cm, bottomMargin=0.7 * cm).build(story)
    paths.append(pdf)
    return paths


def add_to_workbook(kpis: pd.DataFrame, reports: pd.DataFrame) -> Path:
    wb = load_workbook(MASTER)
    for sheet in ["KPI Rationale", "Report Rationale"]:
        if sheet in wb.sheetnames:
            del wb[sheet]
    for sheet, df in [("KPI Rationale", kpis), ("Report Rationale", reports)]:
        ws = wb.create_sheet(sheet)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="111111")
        ws.freeze_panes = "A2"
        for col in "ABCDEFGHIJKL":
            ws.column_dimensions[col].width = 30
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(MASTER)
    return MASTER


def copy_to_repo(paths: list[Path], master: Path) -> None:
    for folder in [REPO / "docs", REPO / "artifacts" / "data_samples", REPO / "artifacts" / "reports"]:
        folder.mkdir(parents=True, exist_ok=True)
    for path in paths:
        if path.suffix.lower() == ".md":
            shutil.copy2(path, REPO / "docs" / path.name)
        elif path.suffix.lower() == ".csv":
            shutil.copy2(path, REPO / "artifacts" / "data_samples" / path.name)
        elif path.suffix.lower() == ".pdf":
            shutil.copy2(path, REPO / "artifacts" / "reports" / path.name)
    shutil.copy2(master, REPO / "artifacts" / "reports" / master.name)


def main() -> None:
    values = load_values()
    kpis = build_kpis(values)
    reports = build_reports()
    paths = write_outputs(kpis, reports)
    master = add_to_workbook(kpis, reports)
    copy_to_repo(paths, master)
    print("KPI rationale layer complete:")
    for path in [*paths, master]:
        print(f"- {path} ({path.stat().st_size / (1024 * 1024):.2f} MB)")


if __name__ == "__main__":
    main()
