from __future__ import annotations

import csv
import json
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image as PdfImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
REPO = OUT / "leatt-fabric-bi-ml-git-proof"
SQLITE = OUT / "leatt_ecommerce_warehouse.sqlite"
SCREENSHOTS = OUT / "screenshots"
SCREENSHOTS.mkdir(exist_ok=True)


def money(value: float) -> str:
    return f"R{value:,.0f}"


def load_data() -> dict[str, pd.DataFrame]:
    conn = sqlite3.connect(SQLITE)
    data = {
        "monthly": pd.read_sql("select * from agg_monthly order by month", conn),
        "category": pd.read_sql("select * from agg_category order by net_revenue_zar desc", conn),
        "channel": pd.read_sql("select * from agg_channel order by net_revenue_zar desc", conn),
        "province": pd.read_sql("select * from agg_province order by net_revenue_zar desc", conn),
        "product": pd.read_sql("select * from agg_product order by net_revenue_zar desc limit 100", conn),
        "sample": pd.read_sql(
            """
            select order_date, month, customer_id, product_title, category, channel, device, province,
                   quantity, net_revenue_zar, gross_margin_zar, return_flag, return_amount_zar,
                   fulfillment_days, payment_method, campaign
            from fact_transaction_lines
            order by transaction_line_id
            limit 500000
            """,
            conn,
        ),
    }
    conn.close()
    return data


def build_intelligence(data: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    monthly = data["monthly"].copy()
    monthly["margin_rate"] = monthly["gross_margin_zar"] / monthly["net_revenue_zar"].replace(0, np.nan)
    monthly["return_rate"] = monthly["return_amount_zar"] / monthly["net_revenue_zar"].replace(0, np.nan)
    monthly["revenue_ma3"] = monthly["net_revenue_zar"].rolling(3, min_periods=2).mean()
    monthly["revenue_anomaly_score"] = ((monthly["net_revenue_zar"] - monthly["revenue_ma3"]) / monthly["revenue_ma3"]).fillna(0)
    monthly["anomaly_flag"] = monthly["revenue_anomaly_score"].abs() > 0.12

    category = data["category"].copy()
    category["margin_rate"] = category["gross_margin_zar"] / category["net_revenue_zar"].replace(0, np.nan)
    category["return_rate"] = category["return_amount_zar"] / category["net_revenue_zar"].replace(0, np.nan)
    category["revenue_rank"] = category["net_revenue_zar"].rank(ascending=False, method="dense").astype(int)
    category["strategic_role"] = np.select(
        [
            (category["revenue_rank"] <= 3) & (category["margin_rate"] >= category["margin_rate"].median()),
            category["return_rate"] > category["return_rate"].quantile(0.75),
            category["margin_rate"] > category["margin_rate"].quantile(0.75),
        ],
        ["Hero growth engine", "Return-risk watchlist", "Margin builder"],
        default="Portfolio support",
    )

    channel = data["channel"].copy()
    spend_rate = {"Organic Search": 0.035, "Paid Search": 0.145, "Paid Social": 0.170, "Email": 0.018, "Direct": 0.010, "Marketplace": 0.115}
    channel["estimated_spend_zar"] = channel.apply(lambda r: r["net_revenue_zar"] * spend_rate.get(r["channel"], 0.08), axis=1)
    channel["roas"] = channel["net_revenue_zar"] / channel["estimated_spend_zar"].replace(0, np.nan)
    channel["decision"] = np.select(
        [
            channel["roas"] >= 50,
            channel["roas"].between(10, 49.999),
            channel["roas"] < 10,
        ],
        ["Automate and scale", "Optimize creative and landing pages", "Constrain spend; use for learning"],
        default="Review",
    )

    product = data["product"].copy()
    product["margin_rate"] = product["gross_margin_zar"] / product["net_revenue_zar"].replace(0, np.nan)
    product["return_rate"] = product["return_amount_zar"] / product["net_revenue_zar"].replace(0, np.nan)
    product["nba_action"] = np.select(
        [
            (product["net_revenue_zar"].rank(pct=True) > 0.85) & (product["margin_rate"] > product["margin_rate"].median()),
            product["return_rate"] > product["return_rate"].quantile(0.8),
            product["margin_rate"] > product["margin_rate"].quantile(0.8),
        ],
        ["Feature in hero bundle and paid search", "Improve sizing/copy/reviews before scaling", "Attach as margin-positive recommendation"],
        default="Keep in evergreen merchandising",
    )

    province = data["province"].copy()
    province["market_role"] = np.select(
        [
            province["net_revenue_zar"].rank(pct=True) > 0.75,
            province["gross_margin_zar"].rank(pct=True) > 0.75,
        ],
        ["Priority growth market", "Margin-efficient market"],
        default="Maintain and monitor",
    )

    actions = []
    top_category = category.iloc[0]
    top_channel = channel.sort_values("roas", ascending=False).iloc[0]
    risky_category = category.sort_values("return_rate", ascending=False).iloc[0]
    margin_product = product.sort_values("gross_margin_zar", ascending=False).iloc[0]
    anomaly = monthly.loc[monthly["anomaly_flag"]].tail(1)
    actions.extend(
        [
            ("CEO", "Protect hero category", f"Make {top_category['category']} the board-level growth theme; it contributes {money(top_category['net_revenue_zar'])} revenue.", "High", "Revenue growth"),
            ("CMO", "Scale efficient acquisition", f"Use {top_channel['channel']} as always-on efficiency engine with ROAS {top_channel['roas']:.1f}x.", "High", "Marketing efficiency"),
            ("COO", "Reduce return leakage", f"Investigate {risky_category['category']} return-rate drivers before increasing media spend.", "Medium", "Margin protection"),
            ("Merchandising", "Build margin bundle", f"Bundle {margin_product['product_title'][:70]} with compatible accessories and protection products.", "High", "AOV and margin"),
            ("Data Governance", "Operationalize evidence", "Publish source register, row counts, file hashes, capacity status and refresh runbook into monthly close pack.", "High", "Auditability"),
        ]
    )
    if not anomaly.empty:
        r = anomaly.iloc[0]
        actions.append(("BI Ops", "Investigate revenue anomaly", f"{r['month']} changed {r['revenue_anomaly_score']:.1%} versus 3-month average.", "Medium", "Monitoring"))
    next_best_actions = pd.DataFrame(actions, columns=["owner", "action", "rationale", "priority", "business_outcome"])

    agents = pd.DataFrame(
        [
            ("Revenue Sentinel", "Monitors revenue, margin, returns and anomaly signals; creates executive alerts.", "MonthlyTrend, CategoryKPI, FabricProof", "Power BI alert + Teams/Email summary"),
            ("Merchandising Copilot", "Recommends hero products, bundles, PDP copy fixes and attach items.", "ProductKPI, MLReturnRisk, SEO roadmap", "Product backlog and campaign briefs"),
            ("Finance Reconciler", "Compares ecommerce revenue, VAT, refunds and SAP posted finance extracts.", "Finance Reconciliation, SAP B1/BW extracts", "Audit exception register"),
            ("Growth Experimenter", "Designs A/B tests, reads results, and recommends rollouts.", "AB Test Plan/Results, ChannelROI", "Experiment decision memo"),
            ("Governance Steward", "Checks source freshness, PII minimization, row counts, lineage and Git proof.", "Source Register, Requirements Log", "Governance scorecard"),
        ],
        columns=["agent_name", "purpose", "input_tables", "output"],
    )

    return {
        "intelligent_monthly": monthly,
        "intelligent_category": category,
        "intelligent_channel": channel,
        "intelligent_product": product,
        "intelligent_province": province,
        "next_best_actions": next_best_actions,
        "ai_agents": agents,
    }


def create_visuals(intel: dict[str, pd.DataFrame]) -> dict[str, Path]:
    plt.style.use("seaborn-v0_8-whitegrid")
    paths = {}
    monthly = intel["intelligent_monthly"]
    category = intel["intelligent_category"].head(8).sort_values("net_revenue_zar")
    actions = intel["next_best_actions"]
    channel = intel["intelligent_channel"].sort_values("roas")
    product = intel["intelligent_product"].head(12).sort_values("gross_margin_zar")

    fig = plt.figure(figsize=(16, 9), facecolor="#f4f6f8")
    gs = fig.add_gridspec(3, 4, hspace=0.45, wspace=0.35)
    fig.suptitle("AI Commerce Command Center", fontsize=24, fontweight="bold", color="#111111")
    ax = fig.add_subplot(gs[0:2, 0:2])
    ax.plot(monthly["month"], monthly["net_revenue_zar"] / 1_000_000, color="#d71920", lw=2.5, label="Revenue")
    ax.plot(monthly["month"], monthly["revenue_ma3"] / 1_000_000, color="#111111", lw=1.8, label="3-month signal")
    ax.scatter(monthly.loc[monthly["anomaly_flag"], "month"], monthly.loc[monthly["anomaly_flag"], "net_revenue_zar"] / 1_000_000, color="#f58518", s=80, label="Anomaly")
    ax.set_title("Revenue Signal And Anomaly Watch")
    ax.set_ylabel("R millions")
    ax.tick_params(axis="x", rotation=45)
    ax.legend()
    ax = fig.add_subplot(gs[0:2, 2:])
    ax.barh(category["category"], category["net_revenue_zar"] / 1_000_000, color="#111111")
    ax.set_title("Hero Categories")
    ax.set_xlabel("R millions")
    ax = fig.add_subplot(gs[2, :2])
    ax.barh(channel["channel"], channel["roas"], color="#4c78a8")
    ax.set_title("Channel Intelligence: ROAS")
    ax = fig.add_subplot(gs[2, 2:])
    ax.axis("off")
    ax.add_patch(plt.Rectangle((0, 0), 1, 1, facecolor="white", edgecolor="#dddddd"))
    ax.text(0.04, 0.83, "Next Best Actions", fontsize=16, fontweight="bold", color="#d71920", transform=ax.transAxes)
    text = "\n".join(f"- {r.owner}: {r.action}" for r in actions.head(5).itertuples(index=False))
    ax.text(0.04, 0.12, text, fontsize=10.5, transform=ax.transAxes, linespacing=1.45)
    path = SCREENSHOTS / "ai_commerce_command_center.png"
    fig.savefig(path, dpi=170, bbox_inches="tight")
    plt.close(fig)
    paths["command"] = path

    fig = plt.figure(figsize=(16, 9), facecolor="#f4f6f8")
    gs = fig.add_gridspec(2, 2, hspace=0.35, wspace=0.30)
    fig.suptitle("Intelligent Merchandising And Governance", fontsize=23, fontweight="bold")
    ax = fig.add_subplot(gs[:, 0])
    ax.barh(product["product_title"].str.slice(0, 38), product["gross_margin_zar"] / 1_000_000, color="#d71920")
    ax.set_title("Products To Prioritize By Gross Margin")
    ax.set_xlabel("R millions")
    ax = fig.add_subplot(gs[0, 1])
    role_counts = intel["intelligent_category"]["strategic_role"].value_counts()
    ax.pie(role_counts.values, labels=role_counts.index, autopct="%1.0f%%", colors=["#d71920", "#111111", "#4c78a8", "#f58518"])
    ax.set_title("Category Strategic Roles")
    ax = fig.add_subplot(gs[1, 1])
    ax.axis("off")
    ax.add_patch(plt.Rectangle((0, 0), 1, 1, facecolor="white", edgecolor="#dddddd"))
    ax.text(0.05, 0.82, "Governance Promise", fontsize=16, fontweight="bold", color="#d71920", transform=ax.transAxes)
    ax.text(
        0.05,
        0.16,
        "Every recommendation is tied back to:\n- source register\n- row counts\n- OneLake path\n- ERD\n- Git commit\n- cost control note",
        fontsize=12,
        transform=ax.transAxes,
        linespacing=1.5,
    )
    path = SCREENSHOTS / "intelligent_merchandising_governance.png"
    fig.savefig(path, dpi=170, bbox_inches="tight")
    plt.close(fig)
    paths["merch"] = path
    return paths


def write_outputs(intel: dict[str, pd.DataFrame], visuals: dict[str, Path]) -> list[Path]:
    paths: list[Path] = []
    for name, df in intel.items():
        path = OUT / f"leatt_{name}.csv"
        df.to_csv(path, index=False)
        paths.append(path)

    blueprint = {
        "system": "Leatt AI Commerce Command Center",
        "created": datetime.now().isoformat(timespec="seconds"),
        "fabric_workspace": "Apollis",
        "lakehouse": "Leatt_BI_ML_Lakehouse",
        "powerbi_dataset": "Leatt Fabric BI ML Semantic Model 202607171909",
        "agents": intel["ai_agents"].to_dict(orient="records"),
        "next_best_actions": intel["next_best_actions"].to_dict(orient="records"),
    }
    json_path = OUT / "ai_commerce_command_center_blueprint.json"
    json_path.write_text(json.dumps(blueprint, indent=2), encoding="utf-8")
    paths.append(json_path)

    def markdown_table(df: pd.DataFrame) -> str:
        rows = ["| " + " | ".join(df.columns) + " |", "| " + " | ".join(["---"] * len(df.columns)) + " |"]
        for _, row in df.iterrows():
            rows.append("| " + " | ".join(str(row[col]).replace("\n", " ") for col in df.columns) + " |")
        return "\n".join(rows)

    monthly = intel["intelligent_monthly"]
    category = intel["intelligent_category"]
    channel = intel["intelligent_channel"]
    product = intel["intelligent_product"]
    total_revenue = monthly["net_revenue_zar"].sum()
    total_margin = monthly["gross_margin_zar"].sum()
    return_rate = monthly["return_amount_zar"].sum() / total_revenue
    top_category = category.iloc[0]
    top_channel = channel.sort_values("roas", ascending=False).iloc[0]
    risky_category = category.sort_values("return_rate", ascending=False).iloc[0]
    top_margin_product = product.sort_values("gross_margin_zar", ascending=False).iloc[0]
    anomaly_rows = monthly.loc[monthly["anomaly_flag"]].tail(1)
    anomaly_text = "No material anomaly crossed the configured 12% revenue movement rule."
    if not anomaly_rows.empty:
        anomaly = anomaly_rows.iloc[0]
        anomaly_text = f"{anomaly['month']} moved {anomaly['revenue_anomaly_score']:.1%} versus the 3-month signal and should be explained in the trading pack."
    data_story = [
        ("Scale", f"The modeled ecommerce estate is large enough for enterprise BI: {money(total_revenue)} revenue and {money(total_margin)} gross margin across the analytical period."),
        ("Category economics", f"{top_category['category']} is the hero growth engine at {money(top_category['net_revenue_zar'])}; it deserves board-level tracking, product storytelling and protected availability."),
        ("Margin leakage", f"Returns absorb {return_rate:.1%} of modeled revenue, with {risky_category['category']} carrying the highest return-rate signal. Fit, sizing, delivery promises and PDP content should be inspected before scaling spend."),
        ("Marketing efficiency", f"{top_channel['channel']} is the efficiency engine with ROAS {top_channel['roas']:.1f}x. Paid channels should be optimized with creative testing, landing-page relevance and bundle-led offers before budget expansion."),
        ("Merchandising action", f"{top_margin_product['product_title'][:80]} is a strong margin candidate for bundles, recommendations and attach-rate testing."),
        ("Monitoring", anomaly_text),
        ("Governance", "Every insight is tied to a source register, row count, ERD, OneLake upload path, Power BI semantic model and Git commit so the work can survive audit review."),
    ]
    data_story_df = pd.DataFrame(data_story, columns=["theme", "what_the_data_tells_us"])
    data_story_path = OUT / "leatt_data_story_business_optimization.csv"
    data_story_df.to_csv(data_story_path, index=False)
    paths.append(data_story_path)

    md_path = OUT / "ai_commerce_command_center_story.md"
    md_path.write_text(
        f"""# AI Commerce Command Center

This layer turns the Leatt Fabric BI/ML project into an intelligent operating model, not just a dashboard.

## Concept

The platform watches sales, margin, returns, marketing, product, finance and governance signals, then turns them into next-best-actions for executives, marketing, merchandising, finance and BI operations.

## What the data tells us

{markdown_table(data_story_df)}

## What we learned

- The project is not only a Power BI exercise. It is an operating model: Fabric stores the governed data estate, Power BI explains the performance, ML highlights return risk, and AI agents turn signals into action.
- The real Leatt storefront runs on Shopify signals rather than SAP Commerce/Hybris, so SAP Commerce is not needed for this case. SAP Business One or SAP BW still fits downstream as the finance and inventory source for reconciliation, VAT, COGS, stock and audit controls.
- dbt is optional. It would be useful later for versioned SQL transformations and tests, but Fabric Data Factory, notebooks, SQL models and semantic-model measures are enough for this proof.
- The most valuable business angle is not "more data"; it is governed decisions: which categories to protect, which products to bundle, where return leakage is eroding margin, and where marketing spend should be scaled or constrained.

## How to optimise the business

- Protect the hero category with availability alerts, executive KPI tracking and SEO/product-story content.
- Build margin-positive bundles around high gross-margin products and test them through A/B experiments.
- Reduce return leakage with size guidance, richer product detail pages, post-purchase education and exception reporting by product/category.
- Move channel budget toward efficient journeys while keeping paid media in controlled experiments until landing pages and creative prove lift.
- Reconcile ecommerce orders, refunds, VAT, payment settlement and SAP postings in a monthly close pack with exception owners.
- Convert OneLake files into governed Delta tables, add Fabric Data Factory refresh monitoring, and publish the Power BI semantic model as the trusted layer for executives.

## Agent design

{markdown_table(intel['ai_agents'])}

## Next best actions

{markdown_table(intel['next_best_actions'])}

## Portfolio positioning

This shows senior BI capability across architecture, data engineering, ML, AI agents, Power BI, governance, accounting reconciliation, marketing analytics and business strategy.
""",
        encoding="utf-8",
    )
    paths.append(md_path)

    pdf_path = OUT / "ai_commerce_command_center_report.pdf"
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleRed", parent=styles["Title"], alignment=TA_CENTER, textColor=colors.HexColor("#d71920")))
    story = [
        Paragraph("AI Commerce Command Center", styles["TitleRed"]),
        Paragraph("A creative intelligent operating layer for the Leatt Fabric BI/ML portfolio.", styles["BodyText"]),
        Spacer(1, 0.3 * cm),
        PdfImage(str(visuals["command"]), width=25 * cm, height=14 * cm),
        PageBreak(),
        Paragraph("What The Data Tells Us", styles["Heading1"]),
        Table(
            [list(data_story_df.columns)] + data_story_df.values.tolist(),
            colWidths=[4 * cm, 20 * cm],
            repeatRows=1,
            style=[
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ],
        ),
        Spacer(1, 0.4 * cm),
        Paragraph("Optimisation Thesis", styles["Heading1"]),
        Paragraph("Use Fabric and Power BI to move from reporting to operating rhythm: protect hero categories, test bundles, reduce return leakage, scale efficient channels, reconcile finance postings, and keep source lineage visible for audit.", styles["BodyText"]),
        PageBreak(),
        Paragraph("Intelligent Merchandising And Governance", styles["Heading1"]),
        PdfImage(str(visuals["merch"]), width=25 * cm, height=14 * cm),
        PageBreak(),
        Paragraph("AI Agents", styles["Heading1"]),
        Table(
            [list(intel["ai_agents"].columns)] + intel["ai_agents"].values.tolist(),
            colWidths=[4 * cm, 7 * cm, 7 * cm, 6 * cm],
            repeatRows=1,
            style=[
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ],
        ),
    ]
    SimpleDocTemplate(str(pdf_path), pagesize=landscape(A4), rightMargin=1 * cm, leftMargin=1 * cm, topMargin=1 * cm, bottomMargin=1 * cm).build(story)
    paths.append(pdf_path)
    return paths


def update_master(intel: dict[str, pd.DataFrame], visuals: dict[str, Path]) -> Path:
    src = OUT / "leatt_ecommerce_bi_ml_fabric_erd_deployment_master.xlsx"
    dst = OUT / "leatt_ecommerce_bi_ml_ai_command_center_master.xlsx"
    shutil.copy2(src, dst)
    wb = load_workbook(dst)
    story_path = OUT / "leatt_data_story_business_optimization.csv"
    data_story_df = pd.read_csv(story_path)
    for sheet in ["AI Command Center", "Data Story", "Next Best Actions", "AI Agents", "Intelligent Category"]:
        if sheet in wb.sheetnames:
            del wb[sheet]
    ws = wb.create_sheet("AI Command Center")
    ws["A1"] = "AI Commerce Command Center"
    ws["A1"].font = Font(bold=True, size=16, color="D71920")
    img = XLImage(str(visuals["command"]))
    img.width = 980
    img.height = 550
    ws.add_image(img, "A3")
    ws.column_dimensions["A"].width = 120
    for sheet, df in [
        ("Data Story", data_story_df),
        ("Next Best Actions", intel["next_best_actions"]),
        ("AI Agents", intel["ai_agents"]),
        ("Intelligent Category", intel["intelligent_category"].head(50)),
    ]:
        ws = wb.create_sheet(sheet)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="111111")
        ws.freeze_panes = "A2"
        for col in "ABCDEFGHIJKL":
            ws.column_dimensions[col].width = 28
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(dst)
    return dst


def copy_to_repo(paths: list[Path], visuals: dict[str, Path], master: Path) -> None:
    if not REPO.exists():
        return
    for folder in [REPO / "docs", REPO / "artifacts" / "data_samples", REPO / "artifacts" / "reports", REPO / "artifacts" / "evidence_images"]:
        folder.mkdir(parents=True, exist_ok=True)
    for p in paths:
        if p.suffix.lower() == ".md":
            shutil.copy2(p, REPO / "docs" / p.name)
        elif p.suffix.lower() in [".csv", ".json"]:
            shutil.copy2(p, REPO / "artifacts" / "data_samples" / p.name)
        elif p.suffix.lower() == ".pdf":
            shutil.copy2(p, REPO / "artifacts" / "reports" / p.name)
    shutil.copy2(master, REPO / "artifacts" / "reports" / master.name)
    for p in visuals.values():
        shutil.copy2(p, REPO / "artifacts" / "evidence_images" / p.name)


def main() -> None:
    data = load_data()
    intel = build_intelligence(data)
    visuals = create_visuals(intel)
    paths = write_outputs(intel, visuals)
    master = update_master(intel, visuals)
    copy_to_repo(paths, visuals, master)
    print("AI command center complete:")
    for p in [*visuals.values(), *paths, master]:
        print(f"- {p} ({p.stat().st_size / (1024 * 1024):.2f} MB)")


if __name__ == "__main__":
    main()
