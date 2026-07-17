from __future__ import annotations

import csv
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
REPO = OUT / "leatt-fabric-bi-ml-git-proof"
SQLITE = OUT / "leatt_ecommerce_warehouse.sqlite"
MASTER = OUT / "leatt_ecommerce_bi_ml_ai_command_center_master.xlsx"


SOURCES = [
    ("Google AI optimization guide", "https://developers.google.com/search/docs/fundamentals/ai-optimization-guide"),
    ("Google ecommerce structured data", "https://developers.google.com/search/docs/specialty/ecommerce/include-structured-data-relevant-to-ecommerce"),
    ("Google Product structured data", "https://developers.google.com/search/docs/appearance/structured-data/product"),
    ("Google ecommerce site structure", "https://developers.google.com/search/docs/specialty/ecommerce/help-google-understand-your-ecommerce-site-structure"),
    ("Google JavaScript rendering guidance", "https://developers.google.com/search/docs/crawling-indexing/javascript/dynamic-rendering"),
    ("Google Merchant Center AI performance insights", "https://support.google.com/merchants/answer/17200695"),
    ("Bing Webmaster Guidelines", "https://www.bing.com/webmasters/help/bing-webmaster-guidelines-30fba23a"),
    ("Bing structured data guidance", "https://www.bing.com/webmasters/help/marking-up-your-site-with-structured-data-3a93e731"),
]


def money(value: float) -> str:
    return f"R{value:,.0f}"


def load_context() -> dict[str, str]:
    conn = sqlite3.connect(SQLITE)
    total_revenue, total_margin, total_returns = conn.execute(
        "select sum(net_revenue_zar), sum(gross_margin_zar), sum(return_amount_zar) from fact_transaction_lines"
    ).fetchone()
    hero_category, hero_revenue = conn.execute(
        "select category, net_revenue_zar from agg_category order by net_revenue_zar desc limit 1"
    ).fetchone()
    conn.close()
    return {
        "total_revenue": money(total_revenue),
        "total_margin": money(total_margin),
        "return_rate": f"{total_returns / total_revenue:.1%}",
        "hero_category": str(hero_category),
        "hero_revenue": money(hero_revenue),
    }


def build_best_practices(ctx: dict[str, str]) -> pd.DataFrame:
    rows = [
        ("AI search foundation", "Treat classic SEO as the foundation for AI Overviews, AI Mode, Bing Copilot and grounding results.", "Google and Bing both state AI search experiences depend on normal crawling, indexing, ranking, retrieval and content quality.", "Keep product/category pages indexable, useful, specific and technically clean; avoid creating a separate 'AI SEO' silo.", "Organic visibility, AI visibility, cited/grounded answer eligibility.", "Search Console, Bing Webmaster Tools, Merchant Center AI performance insights."),
        ("Product structured data", "Implement complete Product/ProductGroup JSON-LD for PDPs and variants.", "Product schema helps search systems understand price, availability, identifiers, variants, shipping and returns.", "Add SKU/GTIN where available, price, availability, images, brand, variants, ratings, shipping and return policy.", "Rich result eligibility, shopping panel coverage, product understanding.", "Structured data validation, rich result eligibility, product data completeness."),
        ("Merchant Center feed", "Use Merchant Center feeds as the trusted product-data layer, not only on-page schema.", "Google says using both structured data and Merchant Center feeds maximizes eligibility and helps verify product data.", "Create/enrich feed attributes for title, description, category, image, price, availability, shipping, returns, GTIN/SKU and product highlights.", "Free listings, shopping surfaces, AI shopping visibility, feed approval health.", "Feed disapprovals, product completeness, free listing impressions, AI share of voice where available."),
        ("Conversational attributes", "Rewrite product/category copy around shopper questions and use-cases, not keyword stuffing.", "AI shopping queries are more conversational and journey-based; Merchant Center AI insights measure query terms and journey phases.", f"For {ctx['hero_category']}, add buying-guide language such as use-case, fit, protection level, terrain, climate, rider type and compatibility.", "AI query coverage, non-brand demand capture, PDP conversion.", "Conversational query coverage, FAQ engagement, Search Console long-query clicks."),
        ("Category hub strategy", "Build authoritative category hubs for hero and high-margin categories.", "Ecommerce navigation and internal links help search engines understand site structure and page importance.", f"Create a {ctx['hero_category']} hub because it contributes {ctx['hero_revenue']} modeled revenue; link to subcategories, guides and PDPs.", "Category rankings, internal PageRank flow, crawl discovery, conversion paths.", "Non-brand category clicks, impressions, indexed category URLs, category revenue."),
        ("Faceted navigation control", "Control filter/facet URLs so useful facets are indexable and low-value combinations do not waste crawl budget.", "Large ecommerce catalogs can generate duplicate/thin URLs; canonical, noindex and sitemap decisions must be deliberate.", "Index SEO-worthy combinations like activity/category; noindex or canonical low-value filters like sort, price-only or temporary parameters.", "Crawl efficiency, duplicate control, category authority.", "Indexed URL count, duplicate title rate, crawl waste, canonical mismatch."),
        ("JavaScript rendering", "Prefer server-side rendering, static rendering or hydration for product/category content.", "Google guidance says dynamic rendering is a workaround, not a long-term solution.", "Ensure product names, prices, images, links, availability and category copy are visible in rendered HTML and linked via href anchors.", "Indexability, crawl reliability, content visibility.", "Rendered HTML checks, Search Console URL inspection, crawl comparison."),
        ("Trust and originality", "Use human-reviewed AI assistance; do not publish scaled auto-generated copy without expertise or value.", "Bing warns against automatically generated low-quality content, keyword stuffing, scraped content and prompt-injection manipulation.", "Use AI to draft briefs, then add expert product knowledge, unique comparisons, fit notes, local shipping/returns and real customer language.", "Trust, ranking resilience, AI citation quality.", "Content QA score, duplicate content, organic assisted conversion, review sentiment."),
        ("Schema plus visible content", "Keep structured data consistent with visible page content.", "Bing and Google can ignore misleading or invalid markup; structured data must represent what users see.", "Validate schema in CI; block deployment when price, availability or return policy markup differs from PDP content/feed.", "Rich result trust, data consistency, reduced eligibility loss.", "Schema errors, merchant feed mismatches, PDP/feed price deltas."),
        ("AI performance measurement", "Create a new AI Search Visibility report, separate from normal SEO but connected to revenue.", "Merchant Center AI performance insights introduces share of voice, competitor average, query frequency and journey phase metrics where available.", "Track discovery/evaluation/purchase query visibility, product terms, competitor presence and content gaps.", "AI share of voice, emerging demand, competitor displacement.", "AI impressions, share of voice, query type, journey phase, competitor average."),
        ("Bing/Copilot readiness", "Submit clean sitemaps and use IndexNow for freshness where appropriate.", "Bing guidelines emphasize sitemaps, freshness, crawlable links and structured data for search and AI grounding.", "Publish canonical sitemap URLs with accurate lastmod; submit important product/category updates via IndexNow.", "Bing/Copilot freshness and grounding eligibility.", "Bing indexed pages, crawl errors, IndexNow submissions, sitemap lastmod accuracy."),
        ("SEO-to-BI loop", "Connect SEO, Merchant Center, GA4, Shopify, Fabric and Power BI into one growth model.", "SEO becomes more valuable when traffic, ranking, product data, returns and margin are evaluated together.", "Join query/category/product data to revenue, margin, returns and stock so SEO recommendations optimize profit, not only clicks.", "Profitable organic growth and reduced wasted content effort.", "Organic revenue, gross margin, return rate by landing page, stockout impact."),
    ]
    return pd.DataFrame(rows, columns=["practice_area", "best_practice", "why_latest_or_relevant", "leatt_action", "what_it_signals", "reporting_kpis"])


def build_ai_seo_kpis() -> pd.DataFrame:
    rows = [
        ("AI Share of Voice", "Share of AI shopping impressions captured versus competitors.", "Signals whether the brand appears in conversational AI shopping journeys.", "Merchant Center AI performance insights where available; supplement with controlled manual/third-party monitoring.", "AI Search Visibility"),
        ("Conversational Query Coverage", "Percent of priority shopper questions covered by category/PDP/guide content.", "Signals whether content matches how people ask AI systems for recommendations.", "Search Console long-tail queries, site search, customer support questions, product review mining.", "Content Gap Report"),
        ("Product Data Completeness", "Percent of products with GTIN/SKU, price, availability, image, shipping, returns, category and rich descriptions.", "Signals eligibility strength for Merchant Center, rich results and AI shopping surfaces.", "Merchant Center feed diagnostics plus schema validation.", "Product Data Quality"),
        ("Structured Data Validity", "Valid Product/ProductGroup/Organization markup rate.", "Signals machine-readable trust and rich result eligibility.", "Rich Results Test, Schema.org validation, crawl extracts.", "Technical SEO / Schema Audit"),
        ("Rich Result Eligibility", "Products/categories eligible for product snippets, merchant listings, images and shopping experiences.", "Signals whether search engines can show enhanced product experiences.", "Google Search Console enhancements, Merchant Center diagnostics.", "Search Appearance Report"),
        ("Index Coverage Health", "Important category/PDP URLs indexed versus submitted canonicals.", "Signals discoverability and crawl/indexation quality.", "Search Console and Bing Webmaster indexed URL reports.", "Technical SEO"),
        ("Crawl Waste Rate", "Low-value parameter/facet URLs crawled as share of total crawl.", "Signals wasted crawl budget and duplicate/thin page risk.", "Crawl logs, sitemap comparison, faceted URL inventory.", "Technical SEO"),
        ("Organic Revenue And Margin", "Revenue and gross margin from organic landing pages.", "Signals whether SEO is generating profitable demand, not just traffic.", "GA4/Shopify/Fabric join by landing page and product/category.", "SEO Revenue Dashboard"),
        ("AI/SEO Assisted Return Rate", "Return rate by organic/AI landing page, category and product.", "Signals whether content is setting correct expectations before purchase.", "Shopify orders, returns, GA4 landing page, Fabric model.", "Returns And Content Quality"),
        ("Content Freshness SLA", "Percent of priority pages refreshed inside target review window.", "Signals whether prices, stock, specs and advice remain trustworthy.", "CMS timestamps, sitemap lastmod, product feed update history.", "Governance Scorecard"),
    ]
    return pd.DataFrame(rows, columns=["kpi", "definition", "what_it_signals", "measurement_source", "report"])


def build_roadmap(ctx: dict[str, str]) -> pd.DataFrame:
    rows = [
        ("0-2 weeks", "Technical audit", "Validate Product/ProductGroup schema, sitemap canonicals, index coverage, rendered HTML and Merchant Center feed health.", "SEO/Data Engineering", "Blockers list with severity, owner and fix path."),
        ("0-2 weeks", "Hero category content sprint", f"Build an AI-search-ready {ctx['hero_category']} hub with buyer questions, use-cases, comparison tables and internal links.", "SEO/Content/Merchandising", "Non-brand category query coverage and richer category authority."),
        ("2-4 weeks", "Product feed enrichment", "Improve titles/descriptions, GTIN/SKU coverage, product highlights, shipping, return policy and high-quality images.", "Ecommerce/Merchant Center", "Higher product data completeness and shopping eligibility."),
        ("2-4 weeks", "Facet governance", "Define index/noindex/canonical rules for size, color, activity, price, sort and campaign parameters.", "SEO/Engineering", "Reduced crawl waste and duplicate URL risk."),
        ("4-6 weeks", "AI visibility dashboard", "Create Power BI report for AI share of voice, conversational query coverage, product term gaps and competitor presence.", "BI/Marketing", "Management visibility into AI-era search demand."),
        ("4-8 weeks", "Profit-based SEO model", "Join SEO landing pages to revenue, margin, stock and returns in Fabric.", "BI/Data Engineering", "SEO priorities ranked by profit and operational risk."),
        ("8-12 weeks", "Experiment loop", "Run A/B tests for category hub copy, PDP fit guidance, bundle messaging and FAQ modules.", "Growth/BI", "Measured lift in conversion, AOV and return reduction."),
    ]
    return pd.DataFrame(rows, columns=["timeframe", "workstream", "action", "owner", "expected_outcome"])


def markdown_table(df: pd.DataFrame) -> str:
    rows = ["| " + " | ".join(df.columns) + " |", "| " + " | ".join(["---"] * len(df.columns)) + " |"]
    for _, row in df.iterrows():
        rows.append("| " + " | ".join(str(row[col]).replace("\n", " ") for col in df.columns) + " |")
    return "\n".join(rows)


def write_outputs(best: pd.DataFrame, kpis: pd.DataFrame, roadmap: pd.DataFrame) -> list[Path]:
    outputs: list[Path] = []
    for name, df in [
        ("leatt_ai_seo_best_practices.csv", best),
        ("leatt_ai_seo_kpis.csv", kpis),
        ("leatt_ai_seo_roadmap.csv", roadmap),
    ]:
        path = OUT / name
        df.to_csv(path, index=False)
        outputs.append(path)

    sources_md = "\n".join(f"- [{name}]({url})" for name, url in SOURCES)
    md = OUT / "ai_seo_best_practices_playbook.md"
    md.write_text(
        f"""# AI SEO Best Practices Playbook

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}

This playbook adds latest AI-era SEO and ecommerce search practices to the Leatt Fabric BI/ML project. The principle is simple: AI search visibility comes from strong technical SEO, trusted product data, structured ecommerce entities, useful content, and measurement that connects visibility to margin.

## Best Practices

{markdown_table(best)}

## AI SEO KPIs

{markdown_table(kpis)}

## 12-Week Roadmap

{markdown_table(roadmap)}

## Recommended Power BI Reports

- AI Search Visibility: share of voice, AI impressions, query type, journey phase, competitor average.
- Product Data Quality: feed completeness, schema validity, GTIN/SKU coverage, shipping/returns coverage.
- Technical SEO: index coverage, crawl waste, canonical mismatch, rendered HTML availability, sitemap freshness.
- Content Gap: conversational questions, category hubs, PDP copy gaps, FAQ coverage, review themes.
- SEO Profitability: organic revenue, gross margin, return rate, stockout impact and conversion by landing page.

## Sources

{sources_md}
""",
        encoding="utf-8",
    )
    outputs.append(md)

    pdf = OUT / "leatt_ai_seo_best_practices_playbook.pdf"
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleRed", parent=styles["Title"], alignment=TA_CENTER, textColor=colors.HexColor("#d71920")))
    pdf_story = [
        Paragraph("AI SEO Best Practices Playbook", styles["TitleRed"]),
        Paragraph("Latest ecommerce SEO and AI search practices mapped to BI signals, KPIs and actions.", styles["BodyText"]),
        Spacer(1, 0.35 * cm),
        Table(
            [list(best.columns)] + best.values.tolist(),
            colWidths=[3.0 * cm, 4.5 * cm, 5.2 * cm, 5.2 * cm, 4.2 * cm, 4.2 * cm],
            repeatRows=1,
            style=[
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
                ("FONTSIZE", (0, 0), (-1, -1), 5.8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ],
        ),
        PageBreak(),
        Paragraph("AI SEO KPI Model", styles["Heading1"]),
        Table(
            [list(kpis.columns)] + kpis.values.tolist(),
            colWidths=[4.0 * cm, 6.0 * cm, 6.0 * cm, 6.0 * cm, 4.0 * cm],
            repeatRows=1,
            style=[
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
                ("FONTSIZE", (0, 0), (-1, -1), 6.8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ],
        ),
        PageBreak(),
        Paragraph("Roadmap", styles["Heading1"]),
        Table(
            [list(roadmap.columns)] + roadmap.values.tolist(),
            colWidths=[3.0 * cm, 4.0 * cm, 8.0 * cm, 4.0 * cm, 7.0 * cm],
            repeatRows=1,
            style=[
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ],
        ),
    ]
    SimpleDocTemplate(str(pdf), pagesize=landscape(A4), rightMargin=0.7 * cm, leftMargin=0.7 * cm, topMargin=0.7 * cm, bottomMargin=0.7 * cm).build(pdf_story)
    outputs.append(pdf)
    return outputs


def update_workbook(best: pd.DataFrame, kpis: pd.DataFrame, roadmap: pd.DataFrame) -> Path:
    wb = load_workbook(MASTER)
    for sheet in ["AI SEO Best Practices", "AI SEO KPIs", "AI SEO Roadmap"]:
        if sheet in wb.sheetnames:
            del wb[sheet]
    for sheet, df in [
        ("AI SEO Best Practices", best),
        ("AI SEO KPIs", kpis),
        ("AI SEO Roadmap", roadmap),
    ]:
        ws = wb.create_sheet(sheet)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="111111")
        ws.freeze_panes = "A2"
        for col in "ABCDEFGHIJKL":
            ws.column_dimensions[col].width = 34
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(MASTER)
    return MASTER


def copy_to_repo(paths: list[Path], master: Path) -> None:
    for folder in [REPO / "docs", REPO / "artifacts" / "data_samples", REPO / "artifacts" / "reports"]:
        folder.mkdir(parents=True, exist_ok=True)
    for path in paths:
        if path.suffix.lower() == ".md":
            shutil.copy2(path, REPO / "docs" / path.name)
        elif path.suffix.lower() == ".csv":
            shutil.copy2(path, REPO / "artifacts" / "data_samples" / path.name)
        elif path.suffix.lower() == ".pdf":
            shutil.copy2(path, REPO / "artifacts" / "reports" / path.name)
    shutil.copy2(master, REPO / "artifacts" / "reports" / master.name)


def main() -> None:
    ctx = load_context()
    best = build_best_practices(ctx)
    kpis = build_ai_seo_kpis()
    roadmap = build_roadmap(ctx)
    paths = write_outputs(best, kpis, roadmap)
    master = update_workbook(best, kpis, roadmap)
    copy_to_repo(paths, master)
    print("AI SEO best practices layer complete:")
    for path in [*paths, master]:
        print(f"- {path} ({path.stat().st_size / (1024 * 1024):.2f} MB)")


if __name__ == "__main__":
    main()
