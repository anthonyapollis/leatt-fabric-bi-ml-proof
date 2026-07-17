from __future__ import annotations

import csv
import shutil
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont
from pypdf import PdfReader, PdfWriter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image as PdfImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
EVIDENCE = OUT / "evidence_images"


SCREENSHOTS = [
    {
        "name": "azure_portal_home_initial",
        "path": Path(r"C:\Users\Anthony.DESKTOP-ES5HL78\AppData\Local\Temp\codex-clipboard-9df8ef02-cd94-462e-93a5-698ebfe08314.png"),
        "caption": "Azure portal opened in the user's signed-in browser. Used as environment proof for Azure/Fabric work context.",
        "redact": True,
    },
    {
        "name": "azure_portal_home_costs",
        "path": Path(r"C:\Users\Anthony.DESKTOP-ES5HL78\AppData\Local\Temp\codex-clipboard-736a1765-c1dd-4408-808e-aa3b96b3e887.png"),
        "caption": "Azure portal home showing active subscription context. Account and cost details redacted for privacy.",
        "redact": True,
    },
    {
        "name": "leatt_about_source",
        "path": Path(r"C:\Users\Anthony.DESKTOP-ES5HL78\AppData\Local\Temp\codex-clipboard-9f3a795f-8a2c-45cc-bbc0-d27e44d4cb66.png"),
        "caption": "Leatt ZA About Us page used for business context and brand narrative.",
        "redact": False,
    },
    {
        "name": "leatt_moto_boots_collection",
        "path": Path(r"C:\Users\Anthony.DESKTOP-ES5HL78\AppData\Local\Temp\codex-clipboard-500bf71f-e442-495d-b5af-f54cdc047549.png"),
        "caption": "Leatt ZA Moto Boots collection page showing product/pricing context from the ecommerce source.",
        "redact": False,
    },
]


SOURCES = [
    {
        "source": "Leatt ZA product JSON",
        "url": "https://za.leatt.com/products.json",
        "type": "Public ecommerce catalog endpoint",
        "used_for": "Product, variant, price, SKU, availability, tags, category classification seed.",
    },
    {
        "source": "Leatt ZA About Us",
        "url": "https://za.leatt.com/pages/about-us",
        "type": "Public website page",
        "used_for": "Brand and business context in the ebook narrative.",
    },
    {
        "source": "Leatt ZA Moto Boots collection",
        "url": "https://za.leatt.com/collections/moto-boots",
        "type": "Public website page and user screenshot",
        "used_for": "Visual proof of product assortment/pricing and portfolio evidence.",
    },
    {
        "source": "Microsoft Fabric Data Factory documentation",
        "url": "https://learn.microsoft.com/en-us/fabric/data-factory/",
        "type": "Official Microsoft documentation",
        "used_for": "Fabric ingestion, transformation, and orchestration process design.",
    },
    {
        "source": "Azure Data Factory documentation",
        "url": "https://learn.microsoft.com/en-us/azure/data-factory/",
        "type": "Official Microsoft documentation",
        "used_for": "Azure ETL/data integration positioning and ADF process evidence.",
    },
    {
        "source": "Azure Data Factory Copy Activity overview",
        "url": "https://learn.microsoft.com/en-us/azure/data-factory/copy-activity-overview",
        "type": "Official Microsoft documentation",
        "used_for": "Copy activity pattern for source-to-cloud movement.",
    },
    {
        "source": "Microsoft Fabric OneLake overview",
        "url": "https://learn.microsoft.com/en-us/fabric/onelake/onelake-overview",
        "type": "Official Microsoft documentation",
        "used_for": "Unified data lake and OneLake architecture explanation.",
    },
    {
        "source": "Microsoft Fabric Lakehouse overview",
        "url": "https://learn.microsoft.com/en-us/fabric/data-engineering/lakehouse-overview",
        "type": "Official Microsoft documentation",
        "used_for": "Lakehouse, Delta, Spark/SQL, Power BI integration design.",
    },
    {
        "source": "Microsoft Fabric subscription / Azure F SKUs",
        "url": "https://learn.microsoft.com/en-us/fabric/enterprise/buy-subscription",
        "type": "Official Microsoft documentation",
        "used_for": "Confirming Fabric capacity can be purchased through Azure portal.",
    },
    {
        "source": "Power BI Direct Lake overview",
        "url": "https://learn.microsoft.com/en-us/fabric/fundamentals/direct-lake-overview",
        "type": "Official Microsoft documentation",
        "used_for": "Power BI semantic model design for high-volume Delta tables in OneLake.",
    },
    {
        "source": "Generated synthetic ecommerce transactions",
        "url": "outputs/leatt_ecommerce_transactions_2m.parquet",
        "type": "Generated dataset",
        "used_for": "Million-row transaction fact table for BI/ML modeling; synthetic, not Leatt actual sales.",
    },
    {
        "source": "Generated SQLite warehouse",
        "url": "outputs/leatt_ecommerce_warehouse.sqlite",
        "type": "Generated local warehouse",
        "used_for": "Dimensional model, fact table, and BI aggregations.",
    },
]


def font(size: int, bold: bool = False):
    candidates = [
        r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\calibrib.ttf" if bold else r"C:\Windows\Fonts\calibri.ttf",
    ]
    for path in candidates:
        if Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def redact_image(src: Path, dst: Path) -> None:
    img = Image.open(src).convert("RGB")
    draw = ImageDraw.Draw(img)
    w, h = img.size
    redactions = [
        (w - 395, 88, w - 10, 160),
        (w - 345, 162, w - 50, 260),
        (1040, 210, min(w - 70, 1295), 250),
        (820, 520, min(w - 75, 1280), 720),
    ]
    for box in redactions:
        x1, y1, x2, y2 = box
        if x1 < w and y1 < h:
            draw.rounded_rectangle((max(0, x1), max(0, y1), min(w, x2), min(h, y2)), radius=8, fill=(245, 245, 245))
            draw.text((max(0, x1) + 8, max(0, y1) + 8), "redacted", fill=(90, 90, 90), font=font(18, True))
    dst.parent.mkdir(parents=True, exist_ok=True)
    img.save(dst)


def prepare_evidence_images() -> list[dict]:
    EVIDENCE.mkdir(parents=True, exist_ok=True)
    prepared = []
    for item in SCREENSHOTS:
        src = item["path"]
        if not src.exists():
            continue
        dst = EVIDENCE / f"{item['name']}.png"
        if item["redact"]:
            redact_image(src, dst)
        else:
            shutil.copy2(src, dst)
        prepared.append({**item, "prepared_path": dst})
    return prepared


def build_pipeline_diagram() -> Path:
    path = EVIDENCE / "fabric_data_factory_pipeline_blueprint.png"
    fig, ax = plt.subplots(figsize=(14, 6))
    ax.axis("off")
    stages = [
        ("Sources", "Leatt product JSON\nGA4 / Ads / Orders\nReturns / Inventory"),
        ("Data Factory", "Copy job / Pipeline\nREST connector\nScheduled orchestration"),
        ("OneLake Bronze", "Raw JSON\nRaw Parquet\nImmutable evidence"),
        ("Lakehouse Silver", "Delta tables\nData quality checks\nConformed dimensions"),
        ("Gold Model", "Star schema\nSemantic model\nDAX measures"),
        ("ML + BI", "Propensity scoring\nForecasting\nPower BI report"),
    ]
    x_positions = [0.06, 0.225, 0.39, 0.555, 0.72, 0.885]
    colors_fill = ["#f3f4f6", "#e8f1ff", "#eef9f1", "#fff7e6", "#f7eefb", "#ffecec"]
    for i, ((title, body), x, fill) in enumerate(zip(stages, x_positions, colors_fill)):
        ax.add_patch(plt.Rectangle((x - 0.065, 0.28), 0.13, 0.42, facecolor=fill, edgecolor="#111111", linewidth=1.3))
        ax.text(x, 0.62, title, ha="center", va="center", fontsize=13, fontweight="bold", color="#111111")
        ax.text(x, 0.43, body, ha="center", va="center", fontsize=10, color="#333333", linespacing=1.3)
        if i < len(stages) - 1:
            ax.annotate("", xy=(x_positions[i + 1] - 0.075, 0.49), xytext=(x + 0.075, 0.49), arrowprops=dict(arrowstyle="->", lw=2, color="#d71920"))
    ax.text(0.5, 0.88, "Azure / Microsoft Fabric Ecommerce BI and ML Process", ha="center", fontsize=18, fontweight="bold")
    ax.text(0.5, 0.14, "Designed for Fabric Data Factory ingestion, OneLake/Lakehouse storage, Direct Lake/Power BI reporting, and Fabric Data Science models.", ha="center", fontsize=11, color="#555555")
    fig.tight_layout()
    fig.savefig(path, dpi=180)
    plt.close(fig)
    return path


def build_source_register_files() -> tuple[Path, Path]:
    csv_path = OUT / "consolidated_source_register.csv"
    md_path = OUT / "consolidated_source_register.md"
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["source", "url", "type", "used_for"])
        writer.writeheader()
        writer.writerows(SOURCES)
    lines = ["# Consolidated Source Register", "", f"Generated: {datetime.now():%Y-%m-%d %H:%M}", ""]
    for row in SOURCES:
        lines.append(f"## {row['source']}")
        lines.append(f"- URL/File: {row['url']}")
        lines.append(f"- Type: {row['type']}")
        lines.append(f"- Used for: {row['used_for']}")
        lines.append("")
    md_path.write_text("\n".join(lines), encoding="utf-8")
    return csv_path, md_path


def add_excel_evidence(prepared_images: list[dict], pipeline_diagram: Path) -> Path:
    src = OUT / "leatt_ecommerce_bi_ml_report.xlsx"
    dst = OUT / "leatt_ecommerce_bi_ml_report_with_evidence.xlsx"
    shutil.copy2(src, dst)
    wb = load_workbook(dst)
    for sheet in ["Source Register", "Evidence Screenshots", "Azure Process Proof"]:
        if sheet in wb.sheetnames:
            del wb[sheet]

    ws = wb.create_sheet("Source Register")
    headers = ["Source", "URL/File", "Type", "Used for"]
    ws.append(headers)
    for row in SOURCES:
        ws.append([row["source"], row["url"], row["type"], row["used_for"]])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111111")
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 72
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 80
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws = wb.create_sheet("Azure Process Proof")
    ws["A1"] = "Azure / Microsoft Fabric Data Factory Process Blueprint"
    ws["A1"].font = Font(bold=True, size=16, color="D71920")
    ws["A2"] = "This diagram documents the process used to frame the project for Fabric Data Factory, OneLake/Lakehouse, semantic modeling, ML scoring, and BI reporting."
    ws["A2"].alignment = Alignment(wrap_text=True)
    img = XLImage(str(pipeline_diagram))
    img.width = 900
    img.height = 385
    ws.add_image(img, "A4")
    ws.column_dimensions["A"].width = 120

    ws = wb.create_sheet("Evidence Screenshots")
    ws["A1"] = "Evidence Screenshots"
    ws["A1"].font = Font(bold=True, size=16, color="D71920")
    ws["A2"] = "Live-user screenshots supplied in the project brief. Azure account/cost areas are redacted."
    ws["A2"].alignment = Alignment(wrap_text=True)
    row = 4
    for item in prepared_images:
        ws[f"A{row}"] = item["caption"]
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)
        image = XLImage(str(item["prepared_path"]))
        image.width = 760
        image.height = 430
        ws.add_image(image, f"A{row + 1}")
        row += 25
    ws.column_dimensions["A"].width = 120
    wb.save(dst)
    return dst


def create_evidence_addendum(prepared_images: list[dict], pipeline_diagram: Path) -> Path:
    path = OUT / "evidence_addendum.pdf"
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleRed", parent=styles["Title"], alignment=TA_CENTER, textColor=colors.HexColor("#d71920")))
    styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=8, leading=10, textColor=colors.HexColor("#555555")))
    story = []
    story.append(Paragraph("Evidence and Source Addendum", styles["TitleRed"]))
    story.append(Paragraph("Consolidated source register, Azure/Fabric process proof, and real project screenshots for portfolio presentation.", styles["BodyText"]))
    story.append(Spacer(1, 0.5 * cm))
    source_rows = [["Source", "Type", "Used for"]] + [[s["source"], s["type"], s["used_for"]] for s in SOURCES]
    story.append(Table(source_rows, colWidths=[4.6 * cm, 4.0 * cm, 8.0 * cm], repeatRows=1, style=[
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(PageBreak())
    story.append(Paragraph("Azure and Fabric Process Proof", styles["Heading1"]))
    story.append(Paragraph("The following blueprint shows the intended Azure/Fabric implementation path: Data Factory ingestion, OneLake bronze storage, Lakehouse Delta modeling, Power BI semantic layer, and ML scoring/forecasting.", styles["BodyText"]))
    story.append(PdfImage(str(pipeline_diagram), width=24 * cm, height=10.2 * cm))
    story.append(PageBreak())

    for item in prepared_images:
        story.append(Paragraph(item["caption"], styles["Heading2"]))
        story.append(PdfImage(str(item["prepared_path"]), width=24 * cm, height=13.5 * cm))
        story.append(Paragraph("Evidence type: user-provided live screenshot. Azure private details redacted where applicable.", styles["Small"]))
        story.append(PageBreak())

    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=1.2 * cm, leftMargin=1.2 * cm, topMargin=1.0 * cm, bottomMargin=1.0 * cm)
    doc.build(story)
    return path


def merge_pdf(addendum: Path) -> Path:
    base = OUT / "leatt_ecommerce_bi_ml_ebook.pdf"
    dst = OUT / "leatt_ecommerce_bi_ml_ebook_with_evidence.pdf"
    writer = PdfWriter()
    for source in [base, addendum]:
        reader = PdfReader(str(source))
        for page in reader.pages:
            writer.add_page(page)
    with dst.open("wb") as fh:
        writer.write(fh)
    return dst


def update_manifest(new_files: list[tuple[str, Path, str]]) -> None:
    manifest_path = OUT / "asset_manifest.csv"
    existing = []
    if manifest_path.exists():
        with manifest_path.open("r", newline="", encoding="utf-8") as fh:
            existing = list(csv.DictReader(fh))
    existing_names = {row["File"] for row in existing}
    for asset, path, description in new_files:
        if path.name not in existing_names:
            existing.append({"Asset": asset, "File": path.name, "Description": description})
    with manifest_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["Asset", "File", "Description"])
        writer.writeheader()
        writer.writerows(existing)


def main() -> None:
    prepared = prepare_evidence_images()
    pipeline = build_pipeline_diagram()
    csv_path, md_path = build_source_register_files()
    excel = add_excel_evidence(prepared, pipeline)
    addendum = create_evidence_addendum(prepared, pipeline)
    pdf = merge_pdf(addendum)
    update_manifest(
        [
            ("Excel BI/ML report with evidence", excel, "Original workbook plus source register, Azure process proof, and imported screenshots."),
            ("PDF ebook with evidence", pdf, "Original ebook plus consolidated source/evidence addendum."),
            ("Evidence addendum", addendum, "Standalone evidence and screenshot PDF."),
            ("Consolidated source register CSV", csv_path, "Machine-readable source register."),
            ("Consolidated source register Markdown", md_path, "Human-readable source register."),
            ("Azure/Fabric process diagram", pipeline, "Generated process blueprint image for portfolio proof."),
        ]
    )
    print("Updated evidence package:")
    for path in [excel, pdf, addendum, csv_path, md_path, pipeline]:
        print(f"- {path}")
    print(f"Imported screenshots: {len(prepared)}")


if __name__ == "__main__":
    main()
