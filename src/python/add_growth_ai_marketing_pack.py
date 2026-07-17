from __future__ import annotations

import csv
import math
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw, ImageFont
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image as PdfImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
EVIDENCE = OUT / "evidence_images"
SQLITE = OUT / "leatt_ecommerce_warehouse.sqlite"
EYEWEAR_SCREENSHOT = Path(r"C:\Users\Anthony.DESKTOP-ES5HL78\AppData\Local\Temp\codex-clipboard-854884a4-ec02-4779-bab8-424031a28fb7.png")


def money(value: float) -> str:
    return f"R{value:,.0f}"


def pct(value: float) -> str:
    return f"{value:.1%}"


def font(size: int, bold: bool = False):
    for path in [r"C:\Windows\Fonts\arialbd.ttf" if bold else r"C:\Windows\Fonts\arial.ttf"]:
        if Path(path).exists():
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


def load_data() -> dict[str, pd.DataFrame]:
    conn = sqlite3.connect(SQLITE)
    data = {
        "monthly": pd.read_sql("select * from agg_monthly", conn),
        "category": pd.read_sql("select * from agg_category", conn),
        "channel": pd.read_sql("select * from agg_channel", conn),
        "province": pd.read_sql("select * from agg_province", conn),
        "product": pd.read_sql("select * from agg_product", conn),
        "eyewear": pd.read_sql(
            """
            select product_title, category, sum(quantity) quantity, sum(net_revenue_zar) net_revenue_zar, sum(gross_margin_zar) gross_margin_zar
            from fact_transaction_lines
            where lower(product_title) like '%sunglasses%' or lower(product_title) like '%goggle%' or lower(product_title) like '%eyewear%'
            group by product_title, category
            order by net_revenue_zar desc
            limit 50
            """,
            conn,
        ),
    }
    conn.close()
    return data


def copy_eyewear_screenshot() -> Path | None:
    if not EYEWEAR_SCREENSHOT.exists():
        return None
    EVIDENCE.mkdir(parents=True, exist_ok=True)
    dst = EVIDENCE / "leatt_lifestyle_eyewear_collection.png"
    shutil.copy2(EYEWEAR_SCREENSHOT, dst)
    return dst


def z_test_two_proportions(convert_a: int, visitors_a: int, convert_b: int, visitors_b: int) -> float:
    p1 = convert_a / visitors_a
    p2 = convert_b / visitors_b
    pooled = (convert_a + convert_b) / (visitors_a + visitors_b)
    se = math.sqrt(pooled * (1 - pooled) * (1 / visitors_a + 1 / visitors_b))
    if se == 0:
        return 1.0
    z = abs((p2 - p1) / se)
    return math.erfc(z / math.sqrt(2))


def make_ab_tests() -> tuple[pd.DataFrame, pd.DataFrame]:
    rng = np.random.default_rng(77)
    plan_rows = [
        ("Eyewear PDP hero", "Lifestyle eyewear page", "Replace generic grid-first experience with rider-use-case hero and lens benefit copy.", "Conversion rate, add-to-cart rate, AOV", "Sunglasses SpeedViz/RideViz/TheViz collection", "14 days", 26000, "High"),
        ("Bundle offer", "Moto + MTB protective gear", "Show helmet + goggle + glove bundle saving at cart and PDP.", "AOV, margin per visitor, attach rate", "Helmet/Goggle/Glove buyers", "21 days", 42000, "Very high"),
        ("Free shipping threshold", "Checkout/cart", "Test R1000 threshold message vs dynamic 'Rxxx away from free shipping' progress bar.", "Cart conversion, shipping abandonment", "All cart sessions", "14 days", 36000, "High"),
        ("Email lifecycle", "CRM", "VIP early-access drop vs standard new-arrivals newsletter.", "Click-through, revenue per email, unsubscribe", "VIP/Retain ML segments", "10 days", 60000, "Very high"),
        ("SEO landing page", "Organic search", "Create 'best MTB goggles South Africa' guide with product cards and comparison copy.", "Organic clicks, assisted revenue, scroll depth", "Non-brand search visitors", "30 days", 18000, "Medium"),
        ("AI recommendation widget", "PDP/cart", "Next-best-category recommendations using ML segment and product affinity.", "Attach rate, margin per visitor", "Returning visitors", "21 days", 30000, "Very high"),
    ]
    plan = pd.DataFrame(plan_rows, columns=["Test", "Surface", "Hypothesis", "Primary KPIs", "Audience", "Duration", "Recommended visitors", "Priority"])

    baseline = np.array([0.029, 0.041, 0.052, 0.018, 0.026, 0.037])
    expected_lift = np.array([0.14, 0.19, 0.10, 0.24, 0.09, 0.17])
    visitors_a = np.array([13000, 21000, 18000, 30000, 9000, 15000])
    visitors_b = visitors_a.copy()
    conversions_a = rng.binomial(visitors_a, baseline)
    conversions_b = rng.binomial(visitors_b, baseline * (1 + expected_lift))
    aov = np.array([1580, 3120, 1980, 2250, 1450, 2420])
    margin_rate = np.array([0.51, 0.48, 0.49, 0.53, 0.50, 0.52])
    results = []
    for i, row in plan.iterrows():
        cr_a = conversions_a[i] / visitors_a[i]
        cr_b = conversions_b[i] / visitors_b[i]
        lift = cr_b / cr_a - 1
        incremental_orders = conversions_b[i] - (visitors_b[i] * cr_a)
        incremental_revenue = incremental_orders * aov[i]
        p_value = z_test_two_proportions(int(conversions_a[i]), int(visitors_a[i]), int(conversions_b[i]), int(visitors_b[i]))
        results.append(
            {
                "Test": row["Test"],
                "Visitors A": int(visitors_a[i]),
                "Conversions A": int(conversions_a[i]),
                "Conversion A": cr_a,
                "Visitors B": int(visitors_b[i]),
                "Conversions B": int(conversions_b[i]),
                "Conversion B": cr_b,
                "Observed lift": lift,
                "P-value": p_value,
                "Statistically significant": "Yes" if p_value < 0.05 else "No",
                "Estimated incremental revenue": incremental_revenue,
                "Estimated incremental gross margin": incremental_revenue * margin_rate[i],
                "Recommendation": "Roll out" if p_value < 0.05 and lift > 0 else "Keep testing",
            }
        )
    return plan, pd.DataFrame(results)


def make_marketing_analysis(data: dict[str, pd.DataFrame]) -> pd.DataFrame:
    channel = data["channel"].copy()
    spend_rate = {
        "Organic Search": 0.035,
        "Paid Search": 0.145,
        "Paid Social": 0.170,
        "Email": 0.018,
        "Direct": 0.010,
        "Marketplace": 0.115,
    }
    channel["estimated_spend_zar"] = channel.apply(lambda r: r["net_revenue_zar"] * spend_rate.get(r["channel"], 0.08), axis=1)
    channel["roas"] = channel["net_revenue_zar"] / channel["estimated_spend_zar"].replace(0, np.nan)
    channel["cac_proxy_zar"] = channel["estimated_spend_zar"] / (channel["quantity"] / 1.4).replace(0, np.nan)
    channel["margin_after_media_zar"] = channel["gross_margin_zar"] - channel["estimated_spend_zar"]
    channel["action"] = channel["channel"].map(
        {
            "Email": "Scale lifecycle automation and VIP early-access drops.",
            "Organic Search": "Invest in SEO hubs and product comparison content.",
            "Paid Search": "Protect high-intent category terms and feed top-margin SKUs.",
            "Paid Social": "Use product demo/UGC creatives; cap spend by margin.",
            "Direct": "Improve returning-user personalization and bundles.",
            "Marketplace": "Use selectively for clearance and demand discovery.",
        }
    )
    return channel.sort_values("margin_after_media_zar", ascending=False)


def make_competitor_matrix() -> pd.DataFrame:
    rows = [
        ("Fox Racing", "MX and MTB gear leader", "Helmets, boots, guards, goggles, gloves, race kits, apparel", "Brand heat, athlete/race culture, category breadth", "Leatt should lean harder into safety science, protection credibility, and South African origin story.", "https://www.foxracing.com/moto.html"),
        ("Alpinestars", "Premium motorsport protection", "MotoGP, motocross, motorcycling, F1 safety gear, boots, gloves, apparel", "Technical authority, motorsport heritage, premium protection halo", "Compete with test evidence, fit guides, and protection-first product storytelling.", "https://www.alpinestars.com/"),
        ("100%", "Performance eyewear and goggles", "Motocross, MTB, offroad goggles, sports sunglasses, helmets", "Eyewear specialization, clarity/anti-fog/ventilation messaging", "For Leatt eyewear, emphasize lens protection, riding conditions, and bundle with helmets.", "https://www.100percent.com/pages/100-goggles"),
        ("Oakley", "Premium sport optics", "Cycling sunglasses, MX goggles, Prizm lens technology", "Optical tech brand equity, virtual try-on, athlete credibility", "Leatt must avoid pure optics battle; position eyewear as protection plus ride-specific value.", "https://www.oakley.com/en-us/category/goggles/motocross"),
        ("SCOTT", "Bike and moto equipment", "MTB/off-road goggles, helmets, bike gear", "Goggle depth, field-of-vision and anti-fog feature messaging", "Use comparison content: Leatt vs Scott by lens, fit, helmet compatibility, price.", "https://www.scott-sports.com/us/en/products/bike-equipment-mw-goggles"),
    ]
    return pd.DataFrame(rows, columns=["Competitor", "Positioning", "Relevant categories", "Strengths", "Leatt opportunity", "Source"])


def make_seo_map() -> pd.DataFrame:
    rows = [
        ("motocross boots south africa", "Transactional", "Moto Boots collection", "Create size/fit guide, compare 4.5/5.5/6.5 boot ranges, add schema FAQ.", "High"),
        ("mtb goggles south africa", "Transactional", "MTB goggles collection", "Comparison hub: lens, fog, fit, helmet compatibility, replacement lenses.", "High"),
        ("best mountain bike sunglasses", "Commercial research", "Lifestyle eyewear collection", "Rank SpeedViz/RideViz/TheViz by riding use case and light condition.", "High"),
        ("motocross neck brace", "Commercial research", "Protection/neck brace PDPs", "Safety-science content: injury mechanism, fitting, helmet compatibility.", "Very high"),
        ("enduro body protection", "Commercial research", "Body protection collection", "Buyer guide by ADV/Enduro/MTB intensity and heat/ventilation need.", "High"),
        ("mtb helmet and goggles bundle", "Transactional", "Bundle landing page", "Bundle builder with compatible helmet/goggle/glove combinations.", "Very high"),
        ("motocross goggles anti fog", "Problem aware", "Goggle PDP/guide", "Technical guide: anti-fog care, tear-off/roll-off use, mud/wet riding.", "Medium"),
        ("adventure motorcycle gear south africa", "Commercial research", "ADV collection", "Long-form ADV pack list and protection hierarchy.", "Medium"),
        ("leatt sunglasses review", "Brand/commercial", "Eyewear PDPs", "Collect reviews, add UGC gallery, structured review schema.", "High"),
        ("mtb knee guards south africa", "Transactional", "Protection collection", "Fit/comfort comparison, climb vs downhill use case.", "High"),
    ]
    return pd.DataFrame(rows, columns=["Keyword/theme", "Intent", "Target page", "Content action", "Priority"])


def make_ai_roadmap() -> pd.DataFrame:
    rows = [
        ("AI product recommender", "Next-best-product and bundle recommendations by rider type, product affinity, and ML segment.", "High", "Medium", "AOV, attach rate, margin per visitor"),
        ("AI search and fit assistant", "Conversational guide for boots, goggles, braces, and protection fit.", "High", "Medium", "Conversion rate, return rate, support deflection"),
        ("Review intelligence", "Summarize reviews/support tickets into quality, sizing, and expectation issues.", "Medium", "Low", "Return rate, product copy improvement backlog"),
        ("Creative testing copilot", "Generate paid social/email variants and tag performance outcomes.", "Medium", "Low", "CTR, CVR, creative fatigue"),
        ("Demand forecasting", "Forecast category demand and peak-season inventory pressure.", "High", "Medium", "Stockout risk, inventory turns, revenue capture"),
        ("SEO content generator with guardrails", "Draft buyer guides, FAQs, comparison pages, and schema markup from approved product data.", "Medium", "Low", "Organic clicks, assisted revenue"),
        ("Fraud/returns anomaly monitor", "Detect unusual return, refund, or payment patterns.", "Medium", "Medium", "Refund leakage, operational risk"),
    ]
    return pd.DataFrame(rows, columns=["AI use case", "Description", "Business value", "Complexity", "Success metric"])


def make_credit_plan() -> pd.DataFrame:
    rows = [
        ("Day 1", "Create Fabric trial/capacity or use Azure F SKU only after approval.", "Confirm workspace, capacity admin, region, budget alert.", "No spend until approved"),
        ("Day 1", "Lakehouse + OneLake landing zone", "Upload parquet, customer scores, catalog CSV, source register.", "Low"),
        ("Day 2", "Data Factory pipeline", "REST product JSON copy, parquet load, scheduled refresh, basic monitoring.", "Low/Medium"),
        ("Day 2", "Warehouse/semantic model", "Create gold star schema and Power BI measures from supplied DAX.", "Low"),
        ("Day 3", "Power BI Direct Lake report", "Executive, product, marketing, A/B, SEO, AI, ML monitor pages.", "Medium"),
        ("Day 3", "Fabric Data Science notebook", "Train propensity and forecast models; register model outputs.", "Medium"),
        ("Ongoing", "Cost control", "Pause capacity when idle, monitor Fabric Capacity Metrics, export cost report.", "Critical"),
    ]
    return pd.DataFrame(rows, columns=["Timing", "Azure/Fabric activity", "Deliverable", "Credit use level"])


def make_campaign_calendar() -> pd.DataFrame:
    rows = [
        ("Week 1", "Eyewear performance push", "SpeedViz/RideViz/TheViz", "Paid social + SEO guide + email", "Lens clarity, eye protection, lifestyle-to-trail crossover"),
        ("Week 2", "Protection science", "Neck brace/body protection", "SEO guide + video demo + retargeting", "Safety credibility and fit confidence"),
        ("Week 3", "Moto boot authority", "Boots 4.5/5.5/6.5", "Search + shopping + comparison landing page", "Fit, stiffness, enduro vs MX choice"),
        ("Week 4", "Bundle builder", "Helmet + goggle + glove", "A/B test + cart recommendation", "Compatible kit, higher AOV, simpler decision"),
        ("Week 5", "VIP early access", "New arrivals/sale", "Email/SMS lifecycle", "Reward loyalty and clear high-intent inventory"),
        ("Week 6", "ADV pack list", "ADV gear", "Long-form SEO + social carousel", "Trip readiness and protection stack"),
    ]
    return pd.DataFrame(rows, columns=["Period", "Campaign", "Product focus", "Channels", "Message"])


def create_charts(ab_results: pd.DataFrame, marketing: pd.DataFrame, seo: pd.DataFrame, competitor: pd.DataFrame) -> dict[str, Path]:
    plt.style.use("seaborn-v0_8-whitegrid")
    paths: dict[str, Path] = {}
    fig, ax = plt.subplots(figsize=(9, 4.8))
    plot_df = ab_results.sort_values("Estimated incremental gross margin")
    ax.barh(plot_df["Test"], plot_df["Estimated incremental gross margin"] / 1000, color="#d71920")
    ax.set_title("A/B Test Incremental Gross Margin Opportunity")
    ax.set_xlabel("R thousands")
    fig.tight_layout()
    p = EVIDENCE / "ab_test_margin_opportunity.png"
    fig.savefig(p, dpi=180)
    paths["ab"] = p
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(9, 4.8))
    plot_df = marketing.sort_values("roas")
    ax.barh(plot_df["channel"], plot_df["roas"], color="#111111")
    ax.set_title("Estimated Marketing ROAS by Channel")
    ax.set_xlabel("Revenue / estimated media spend")
    fig.tight_layout()
    p = EVIDENCE / "marketing_roas.png"
    fig.savefig(p, dpi=180)
    paths["marketing"] = p
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(8, 4.6))
    counts = seo["Priority"].value_counts().reindex(["Very high", "High", "Medium"], fill_value=0)
    ax.pie(counts.values, labels=counts.index, autopct="%1.0f%%", colors=["#d71920", "#f58518", "#4c78a8"])
    ax.set_title("SEO Opportunity Priority Mix")
    fig.tight_layout()
    p = EVIDENCE / "seo_priority_mix.png"
    fig.savefig(p, dpi=180)
    paths["seo"] = p
    plt.close(fig)

    fig, ax = plt.subplots(figsize=(10, 4.8))
    ax.axis("off")
    ax.text(0.5, 0.92, "Competitive Positioning: Where Leatt Can Win", ha="center", fontsize=17, fontweight="bold")
    cards = [
        ("Safety Science", "Own the protection story and neck-brace heritage."),
        ("Ride-Specific Fit", "Guides by Moto, MTB, ADV, Enduro, and lifestyle."),
        ("Bundle Intelligence", "Helmet + eyewear + gloves + guards compatibility."),
        ("Local Relevance", "South African origin and ZA ecommerce experience."),
    ]
    for i, (title, body) in enumerate(cards):
        x = 0.08 + i * 0.23
        ax.add_patch(plt.Rectangle((x, 0.25), 0.18, 0.42, facecolor=["#ffecec", "#eef9f1", "#e8f1ff", "#fff7e6"][i], edgecolor="#111111"))
        ax.text(x + 0.09, 0.57, title, ha="center", va="center", fontsize=12, fontweight="bold")
        ax.text(x + 0.09, 0.40, body, ha="center", va="center", fontsize=9, wrap=True)
    fig.tight_layout()
    p = EVIDENCE / "competitor_positioning.png"
    fig.savefig(p, dpi=180)
    paths["competitor"] = p
    plt.close(fig)
    return paths


def write_excel(
    data: dict[str, pd.DataFrame],
    plan: pd.DataFrame,
    ab_results: pd.DataFrame,
    marketing: pd.DataFrame,
    competitor: pd.DataFrame,
    seo: pd.DataFrame,
    ai: pd.DataFrame,
    credit: pd.DataFrame,
    calendar: pd.DataFrame,
    charts: dict[str, Path],
    eyewear_image: Path | None,
) -> Path:
    path = OUT / "leatt_growth_ai_marketing_pack.xlsx"
    with pd.ExcelWriter(path, engine="xlsxwriter") as writer:
        summary = pd.DataFrame(
            [
                ("Top A/B rollout", ab_results.sort_values("Estimated incremental gross margin", ascending=False).iloc[0]["Test"]),
                ("Best estimated ROAS channel", marketing.sort_values("roas", ascending=False).iloc[0]["channel"]),
                ("Highest revenue category", data["category"].sort_values("net_revenue_zar", ascending=False).iloc[0]["category"]),
                ("Competitors analyzed", len(competitor)),
                ("SEO themes mapped", len(seo)),
                ("AI use cases prioritized", len(ai)),
                ("Azure credit guidance", "Use Fabric only with approval; pause capacity when idle."),
            ],
            columns=["Item", "Value"],
        )
        summary.to_excel(writer, sheet_name="Growth Executive View", index=False, startrow=2)
        plan.to_excel(writer, sheet_name="AB Test Plan", index=False)
        ab_results.to_excel(writer, sheet_name="AB Test Results", index=False)
        marketing.to_excel(writer, sheet_name="Marketing ROI", index=False)
        competitor.to_excel(writer, sheet_name="Competitor Analysis", index=False)
        seo.to_excel(writer, sheet_name="SEO Roadmap", index=False)
        ai.to_excel(writer, sheet_name="AI Roadmap", index=False)
        credit.to_excel(writer, sheet_name="Azure Credit Plan", index=False)
        calendar.to_excel(writer, sheet_name="Campaign Calendar", index=False)
        data["eyewear"].to_excel(writer, sheet_name="Eyewear Product Focus", index=False)
        wb = writer.book
        title = wb.add_format({"bold": True, "font_size": 18, "font_color": "#d71920"})
        header = wb.add_format({"bold": True, "font_color": "white", "bg_color": "#111111", "border": 1})
        money_fmt = wb.add_format({"num_format": "R#,##0"})
        pct_fmt = wb.add_format({"num_format": "0.0%"})
        for ws in writer.sheets.values():
            ws.freeze_panes(1, 0)
            ws.set_row(0, None, header)
            ws.set_column(0, 0, 28)
            ws.set_column(1, 12, 24)
        ws = writer.sheets["Growth Executive View"]
        ws.write(0, 0, "Leatt Growth, AI, A/B Testing, Marketing and SEO Pack", title)
        ws.write(1, 0, "Built from catalog/transaction model, competitor research, SEO strategy and simulated experiment evidence.")
        ws.insert_image("D3", str(charts["ab"]), {"x_scale": 0.62, "y_scale": 0.62})
        ws.insert_image("D21", str(charts["marketing"]), {"x_scale": 0.62, "y_scale": 0.62})
        if eyewear_image:
            ws.insert_image("A12", str(eyewear_image), {"x_scale": 0.36, "y_scale": 0.36})
        writer.sheets["AB Test Results"].set_column(8, 11, 18, money_fmt)
        writer.sheets["AB Test Results"].set_column(3, 7, 14, pct_fmt)
        writer.sheets["Marketing ROI"].set_column(2, 7, 18, money_fmt)
        writer.sheets["Marketing ROI"].set_column(8, 8, 12)
    return path


def add_sheets_to_master(
    plan: pd.DataFrame,
    ab_results: pd.DataFrame,
    marketing: pd.DataFrame,
    competitor: pd.DataFrame,
    seo: pd.DataFrame,
    ai: pd.DataFrame,
    credit: pd.DataFrame,
    calendar: pd.DataFrame,
    charts: dict[str, Path],
    eyewear_image: Path | None,
) -> Path:
    src = OUT / "leatt_ecommerce_bi_ml_report_with_evidence.xlsx"
    dst = OUT / "leatt_ecommerce_bi_ml_report_full_growth_evidence.xlsx"
    shutil.copy2(src, dst)
    wb = load_workbook(dst)
    dfs = {
        "Growth Strategy": pd.DataFrame(
            [
                ["Portfolio objective", "Show BI, ML, experimentation, marketing analytics, SEO, competitor analysis, and AI strategy in one case study."],
                ["Credit note", "Azure credits should be used only after approval; recommended path is Fabric trial/capacity, Lakehouse, Data Factory, Direct Lake report, then pause."],
                ["Best A/B bet", ab_results.sort_values("Estimated incremental gross margin", ascending=False).iloc[0]["Test"]],
                ["Best channel bet", marketing.sort_values("roas", ascending=False).iloc[0]["channel"]],
            ],
            columns=["Item", "Recommendation"],
        ),
        "AB Testing": ab_results,
        "Marketing ROI": marketing,
        "Competitor Analysis": competitor,
        "SEO Roadmap": seo,
        "AI Roadmap": ai,
        "Azure Credit Plan": credit,
        "Campaign Calendar": calendar,
    }
    for sheet, df in dfs.items():
        if sheet in wb.sheetnames:
            del wb[sheet]
        ws = wb.create_sheet(sheet)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="111111")
        ws.freeze_panes = "A2"
        for col in "ABCDEFGHIJKL":
            ws.column_dimensions[col].width = 24
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws = wb["Growth Strategy"]
    ws["D1"] = "Growth Evidence Visuals"
    ws["D1"].font = Font(bold=True, size=14, color="D71920")
    img = XLImage(str(charts["ab"]))
    img.width = 520
    img.height = 290
    ws.add_image(img, "D3")
    img = XLImage(str(charts["marketing"]))
    img.width = 520
    img.height = 290
    ws.add_image(img, "D20")
    if eyewear_image:
        img = XLImage(str(eyewear_image))
        img.width = 520
        img.height = 290
        ws.add_image(img, "D37")
    wb.save(dst)
    return dst


def write_pdf(
    plan: pd.DataFrame,
    ab_results: pd.DataFrame,
    marketing: pd.DataFrame,
    competitor: pd.DataFrame,
    seo: pd.DataFrame,
    ai: pd.DataFrame,
    credit: pd.DataFrame,
    calendar: pd.DataFrame,
    charts: dict[str, Path],
    eyewear_image: Path | None,
) -> Path:
    path = OUT / "leatt_growth_ai_marketing_report.pdf"
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleRed", parent=styles["Title"], alignment=TA_CENTER, textColor=colors.HexColor("#d71920")))
    styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=8, leading=10, textColor=colors.HexColor("#555555")))
    story = [
        Paragraph("Leatt Growth, AI, Marketing and Experimentation Report", styles["TitleRed"]),
        Paragraph("A/B testing, marketing ROI, competitor analysis, SEO roadmap, AI initiatives, and Azure/Fabric credit-use plan.", styles["BodyText"]),
        Spacer(1, 0.4 * cm),
    ]
    if eyewear_image:
        story.append(PdfImage(str(eyewear_image), width=24 * cm, height=13.5 * cm))
        story.append(Paragraph("Real Leatt ZA eyewear collection screenshot supplied in the project brief.", styles["Small"]))
        story.append(PageBreak())
    story.append(Paragraph("Executive Growth Plays", styles["Heading1"]))
    plays = [
        f"Roll out the strongest A/B winner first: {ab_results.sort_values('Estimated incremental gross margin', ascending=False).iloc[0]['Test']}.",
        f"Use {marketing.sort_values('roas', ascending=False).iloc[0]['channel']} as the efficiency engine, then use paid search/social for controlled scaling.",
        "Compete against Fox, Alpinestars, 100%, Oakley and SCOTT by owning safety science, fit confidence, and smart bundles.",
        "Build SEO hubs around motocross boots, MTB goggles, neck braces, body protection, and eyewear comparison searches.",
        "Use Azure/Fabric credits for a short, controlled proof: Lakehouse + Data Factory + Direct Lake report + Data Science notebook, then pause capacity.",
    ]
    for item in plays:
        story.append(Paragraph(item, styles["BodyText"]))
        story.append(Spacer(1, 0.15 * cm))
    story.append(PdfImage(str(charts["ab"]), width=22 * cm, height=11.5 * cm))
    story.append(PageBreak())

    story.append(Paragraph("A/B Testing Results", styles["Heading1"]))
    table = [["Test", "Lift", "P-value", "Inc. margin", "Decision"]] + [
        [r["Test"], pct(r["Observed lift"]), f"{r['P-value']:.3f}", money(r["Estimated incremental gross margin"]), r["Recommendation"]]
        for _, r in ab_results.iterrows()
    ]
    story.append(Table(table, colWidths=[6.2 * cm, 2.4 * cm, 2.4 * cm, 3.2 * cm, 3.0 * cm], repeatRows=1, style=[
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(PageBreak())

    story.append(Paragraph("Marketing ROI", styles["Heading1"]))
    story.append(PdfImage(str(charts["marketing"]), width=22 * cm, height=11.5 * cm))
    table = [["Channel", "Revenue", "Spend", "ROAS", "Action"]] + [
        [r["channel"], money(r["net_revenue_zar"]), money(r["estimated_spend_zar"]), f"{r['roas']:.1f}x", r["action"]]
        for _, r in marketing.iterrows()
    ]
    story.append(Table(table, colWidths=[3 * cm, 3 * cm, 3 * cm, 2 * cm, 7 * cm], repeatRows=1, style=[
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
    ]))
    story.append(PageBreak())

    story.append(Paragraph("Competitor and SEO Strategy", styles["Heading1"]))
    story.append(PdfImage(str(charts["competitor"]), width=22 * cm, height=10.5 * cm))
    comp_table = [["Competitor", "Strength", "Leatt opportunity"]] + [[r["Competitor"], r["Strengths"], r["Leatt opportunity"]] for _, r in competitor.iterrows()]
    story.append(Table(comp_table, colWidths=[3.3 * cm, 6.2 * cm, 8.4 * cm], repeatRows=1, style=[
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(PageBreak())
    story.append(Paragraph("SEO Roadmap", styles["Heading1"]))
    seo_table = [["Keyword/theme", "Intent", "Target page", "Priority"]] + [[r["Keyword/theme"], r["Intent"], r["Target page"], r["Priority"]] for _, r in seo.iterrows()]
    story.append(Table(seo_table, colWidths=[5.8 * cm, 3.2 * cm, 5.3 * cm, 2.3 * cm], repeatRows=1, style=[
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
    ]))
    story.append(PageBreak())
    story.append(Paragraph("AI and Azure Credit Plan", styles["Heading1"]))
    ai_table = [["AI use case", "Value", "Complexity", "Metric"]] + [[r["AI use case"], r["Business value"], r["Complexity"], r["Success metric"]] for _, r in ai.iterrows()]
    story.append(Table(ai_table, colWidths=[5 * cm, 3 * cm, 3 * cm, 6 * cm], repeatRows=1, style=[
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
    ]))
    story.append(Spacer(1, 0.4 * cm))
    credit_table = [["Timing", "Azure/Fabric activity", "Deliverable", "Credit use"]] + [[r["Timing"], r["Azure/Fabric activity"], r["Deliverable"], r["Credit use level"]] for _, r in credit.iterrows()]
    story.append(Table(credit_table, colWidths=[2.2 * cm, 5.4 * cm, 7.2 * cm, 3 * cm], repeatRows=1, style=[
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
    ]))
    story.append(PageBreak())
    story.append(Paragraph("Six Week Campaign Calendar", styles["Heading1"]))
    cal_table = [["Period", "Campaign", "Focus", "Channels", "Message"]] + [[r["Period"], r["Campaign"], r["Product focus"], r["Channels"], r["Message"]] for _, r in calendar.iterrows()]
    story.append(Table(cal_table, colWidths=[2 * cm, 4 * cm, 3.2 * cm, 4 * cm, 5 * cm], repeatRows=1, style=[
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=1.2 * cm, leftMargin=1.2 * cm, topMargin=1 * cm, bottomMargin=1 * cm)
    doc.build(story)
    return path


def write_csvs(**frames: pd.DataFrame) -> list[Path]:
    paths = []
    for name, df in frames.items():
        p = OUT / f"leatt_{name}.csv"
        df.to_csv(p, index=False)
        paths.append(p)
    return paths


def update_source_register() -> None:
    csv_path = OUT / "consolidated_source_register.csv"
    rows = []
    if csv_path.exists():
        with csv_path.open("r", newline="", encoding="utf-8") as fh:
            rows = list(csv.DictReader(fh))
    additions = [
        ("Fox Racing Moto Gear", "https://www.foxracing.com/moto.html", "Competitor official site", "Competitor category and positioning analysis."),
        ("Fox Racing MTB Gear", "https://www.foxracing.com/mtb.html", "Competitor official site", "Competitor MTB category analysis."),
        ("100% Goggles", "https://www.100percent.com/pages/100-goggles", "Competitor official site", "Eyewear/goggle positioning analysis."),
        ("Oakley MX Goggles", "https://www.oakley.com/en-us/category/goggles/motocross", "Competitor official site", "Optics and motocross eyewear positioning."),
        ("SCOTT MTB Goggles", "https://www.scott-sports.com/us/en/products/bike-equipment-mw-goggles", "Competitor official site", "Goggle product/pricing range comparison."),
        ("Leatt eyewear screenshot", "outputs/evidence_images/leatt_lifestyle_eyewear_collection.png", "User-provided screenshot", "Evidence for eyewear product and pricing context."),
    ]
    existing = {r.get("url") for r in rows}
    for source, url, typ, used_for in additions:
        if url not in existing:
            rows.append({"source": source, "url": url, "type": typ, "used_for": used_for})
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["source", "url", "type", "used_for"])
        writer.writeheader()
        writer.writerows(rows)
    md_path = OUT / "consolidated_source_register.md"
    lines = ["# Consolidated Source Register", "", f"Updated: {datetime.now():%Y-%m-%d %H:%M}", ""]
    for row in rows:
        lines.extend([f"## {row['source']}", f"- URL/File: {row['url']}", f"- Type: {row['type']}", f"- Used for: {row['used_for']}", ""])
    md_path.write_text("\n".join(lines), encoding="utf-8")


def update_manifest(paths: list[Path]) -> None:
    manifest = OUT / "asset_manifest.csv"
    rows = []
    if manifest.exists():
        with manifest.open("r", newline="", encoding="utf-8") as fh:
            rows = list(csv.DictReader(fh))
    existing = {r["File"] for r in rows}
    descriptions = {
        "leatt_growth_ai_marketing_pack.xlsx": "Growth workbook covering A/B testing, marketing ROI, competitor analysis, SEO, AI and Azure credit plan.",
        "leatt_growth_ai_marketing_report.pdf": "Portfolio PDF for growth, marketing, SEO, competitor and AI strategy.",
        "leatt_ecommerce_bi_ml_report_full_growth_evidence.xlsx": "Master Excel report with BI, ML, evidence, growth, SEO, AI and A/B testing sheets.",
    }
    for p in paths:
        if p.name not in existing:
            rows.append({"Asset": p.stem, "File": p.name, "Description": descriptions.get(p.name, "Growth/marketing/AI output asset.")})
    with manifest.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=["Asset", "File", "Description"])
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    EVIDENCE.mkdir(parents=True, exist_ok=True)
    data = load_data()
    eyewear = copy_eyewear_screenshot()
    plan, ab_results = make_ab_tests()
    marketing = make_marketing_analysis(data)
    competitor = make_competitor_matrix()
    seo = make_seo_map()
    ai = make_ai_roadmap()
    credit = make_credit_plan()
    calendar = make_campaign_calendar()
    charts = create_charts(ab_results, marketing, seo, competitor)
    excel = write_excel(data, plan, ab_results, marketing, competitor, seo, ai, credit, calendar, charts, eyewear)
    master = add_sheets_to_master(plan, ab_results, marketing, competitor, seo, ai, credit, calendar, charts, eyewear)
    pdf = write_pdf(plan, ab_results, marketing, competitor, seo, ai, credit, calendar, charts, eyewear)
    csvs = write_csvs(
        ab_test_plan=plan,
        ab_test_results=ab_results,
        marketing_roi=marketing,
        competitor_analysis=competitor,
        seo_roadmap=seo,
        ai_roadmap=ai,
        azure_credit_plan=credit,
        campaign_calendar=calendar,
    )
    update_source_register()
    update_manifest([excel, master, pdf, *csvs, *charts.values(), *( [eyewear] if eyewear else [] )])
    print("Growth pack complete:")
    for p in [excel, master, pdf, *csvs]:
        print(f"- {p} ({p.stat().st_size / (1024 * 1024):.2f} MB)")


if __name__ == "__main__":
    main()
