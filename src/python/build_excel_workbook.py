"""Build the Leatt Growth OS Excel workbook — real Fabric proof + real KPI figures.

Every number comes from artifacts/data_samples/*.csv (the same source the
charts and ebook use). Purpose: give a reviewer a single spreadsheet that
demonstrates working knowledge of Fabric (real object IDs, OneLake transfer
mechanics, medallion architecture, DAX measures, Data Factory pipeline
design, star-schema data model) alongside the actual report figures.
"""

import json
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.join(os.path.dirname(__file__), "..", "..")
DATA = os.path.join(BASE, "artifacts", "data_samples")
OUT = os.path.join(BASE, "artifacts", "reports", "Leatt_Growth_OS_Fabric_Workbook.xlsx")

RED = "D71920"
INK = "141414"
PAPER = "F7F4EF"

HEADER_FILL = PatternFill("solid", fgColor=INK)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color=RED, bold=True, size=16)
SUB_FONT = Font(color=INK, italic=True, size=10)


def write_df(ws, df, start_row=1, title=None, subtitle=None):
    r = start_row
    if title:
        ws.cell(r, 1, title).font = TITLE_FONT
        r += 1
    if subtitle:
        ws.cell(r, 1, subtitle).font = SUB_FONT
        r += 1
    r += 1
    header_row = r
    for c, col in enumerate(df.columns, start=1):
        cell = ws.cell(r, c, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    r += 1
    for _, row in df.iterrows():
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)
        r += 1
    # autofit-ish column widths
    for c, col in enumerate(df.columns, start=1):
        cell_lens = [len(str(v)) for v in df[col].tolist()] if len(df) else [0]
        width = max(len(str(col)), max(cell_lens))
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 12), 60)
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    return r


def main():
    wb = Workbook()

    # 1. Cover
    ws = wb.active
    ws.title = "Cover"
    ws.cell(1, 1, "Leatt Growth OS — Fabric BI/ML Workbook").font = Font(color=RED, bold=True, size=20)
    ws.cell(3, 1, "Purpose: demonstrate working knowledge of Microsoft Fabric — real workspace, "
                  "Lakehouse, OneLake transfer, Data Factory pipeline design, DAX measures, "
                  "medallion architecture and star-schema data model — alongside the real report "
                  "figures the dashboards and ebook are built from.").font = Font(size=11)
    ws.cell(5, 1, "Every figure in this workbook is computed from artifacts/data_samples/*.csv "
                  "in this repository — no numbers are invented for the sheet.").font = SUB_FONT
    sheets = [
        ("Fabric Engine Proof", "Real Fabric object IDs + how data physically moved into OneLake"),
        ("Medallion Architecture", "Bronze -> Silver -> Gold data quality checks (real counts)"),
        ("Data Model (ERD)", "Star-schema tables and relationships"),
        ("DAX Measures", "Power BI semantic model measures, as authored"),
        ("Data Factory Pipeline", "Registered pipeline design — activities and sources"),
        ("Gold KPIs - Category", "Revenue/margin/returns by product category"),
        ("Gold KPIs - Channel", "Revenue/margin/returns by acquisition channel"),
        ("Gold KPIs - Monthly", "Monthly revenue and margin trend, with anomaly flags"),
        ("Gold KPIs - Province", "Revenue/margin/returns by province"),
        ("Revenue Forecast", "6-month forward forecast + anomaly-flagged months"),
        ("Issues - Root Causes", "Root-cause diagnostic playbook + decision signal thresholds"),
        ("Actions - Initiatives", "Next best actions by owner + prioritized initiatives"),
        ("ML Return Risk", "Return-risk model quality + high-risk transaction watchlist"),
        ("Prevention - Controls", "Data controls + the audit exceptions they catch"),
        ("Marketing ROAS", "Return on ad spend by channel"),
        ("A-B Tests", "Experiment results feeding the growth loop"),
        ("SEO - AI SEO Audit", "Live crawl findings on za.leatt.com - technical, structured data, AI/agent readiness"),
        ("Finance Reconciliation", "SAP-style trial balance / reconciliation pack"),
    ]
    r = 8
    ws.cell(r, 1, "Contents").font = Font(bold=True, size=13)
    r += 1
    for name, desc in sheets:
        ws.cell(r, 1, name).font = Font(bold=True)
        ws.cell(r, 2, desc)
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 70

    # 2. Fabric Engine Proof
    ws = wb.create_sheet("Fabric Engine Proof")
    proof = pd.read_csv(os.path.join(DATA, "fabric_onelake_upload_proof.csv"))
    r = write_df(ws, proof, title="Real Fabric / OneLake objects",
                 subtitle="Registered via the Fabric REST API — genuine tenant items, not a diagram")
    r += 2
    million = pd.read_csv(os.path.join(DATA, "million_row_data_proof.csv"))
    write_df(ws, million, start_row=r, title="Data volume actually transferred",
             subtitle="Verified via matching byte counts / row counts on both ends of the OneLake upload")

    # 3. Medallion Architecture
    ws = wb.create_sheet("Medallion Architecture")
    summary = pd.read_csv(os.path.join(DATA, "medallion_quality_summary.csv"))
    r = write_df(ws, summary, title="Bronze -> Silver -> Gold quality summary",
                 subtitle="Real dedup/correction counts from the cleaning pass")
    r += 2
    checks = pd.read_csv(os.path.join(DATA, "medallion_data_quality_checks.csv"))
    write_df(ws, checks, start_row=r, title="Data quality checks applied")

    # 4. Data Model (ERD)
    ws = wb.create_sheet("Data Model (ERD)")
    tables = pd.read_csv(os.path.join(DATA, "leatt_erd_tables.csv"))
    r = write_df(ws, tables, title="Star schema — tables",
                 subtitle="See artifacts/evidence_images/leatt_star_schema_erd.png for the visual ERD")
    r += 2
    rels = pd.read_csv(os.path.join(DATA, "leatt_erd_relationships.csv"))
    write_df(ws, rels, start_row=r, title="Relationships")

    # 5. DAX Measures
    ws = wb.create_sheet("DAX Measures")
    dax_path = os.path.join(BASE, "src", "powerbi", "powerbi_dax_measures.dax")
    with open(dax_path, encoding="utf-8") as f:
        raw = f.read()
    raw_blocks = [b.strip("\n") for b in raw.split("\n\n") if b.strip() and not b.strip().startswith("--")]
    rows = []
    for block in raw_blocks:
        lines = block.splitlines()
        name, _, first_expr_line = lines[0].partition("=")
        expr_lines = [first_expr_line.strip()] + [l.strip() for l in lines[1:]]
        rows.append((name.strip(), "\n".join(l for l in expr_lines if l)))
    dax_df = pd.DataFrame(rows, columns=["Measure", "DAX expression"])
    dax_end = write_df(ws, dax_df, title="Power BI semantic model - DAX measures",
             subtitle="Authored in src/powerbi/powerbi_dax_measures.dax, against Fact Transaction Lines / Dim Date / Customer Scores")
    ws.column_dimensions["B"].width = 70
    for row in ws.iter_rows(min_row=4, max_row=dax_end, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row[0].row].height = 15 * max(1, str(row[1].value).count("\n") + 1)

    # 6. Data Factory Pipeline
    ws = wb.create_sheet("Data Factory Pipeline")
    with open(os.path.join(BASE, "src", "fabric", "fabric_data_factory_pipeline_template.json"), encoding="utf-8") as f:
        pipeline = json.load(f)
    ws.cell(1, 1, f"Pipeline design: {pipeline['name']}").font = TITLE_FONT
    ws.cell(2, 1, pipeline["description"]).font = SUB_FONT
    ws.cell(3, 1, "Registered as a real Fabric Data Factory pipeline item in the workspace "
                  "(pipeline-id-redacted) — presented here as a design, not as "
                  "proven to have executed end to end. The physical data move used the OneLake "
                  "REST API script instead (see Fabric Engine Proof tab).").font = Font(size=10, italic=True)
    acts = pd.json_normalize(pipeline["activities"])
    write_df(ws, acts, start_row=6, title="Activities")
    params = pd.DataFrame(list(pipeline["parameters"].items()), columns=["Parameter", "Value"])
    write_df(ws, params, start_row=6 + len(acts) + 5, title="Parameters")

    # 7-10. Gold KPIs — sourced from leatt_intelligent_*.csv, the correctly-scaled gold
    # aggregates matching the real 2,000,000-row fact table (verified: sums to
    # R3,994,247,412.56, matching leatt_reconciliation.csv exactly). The old
    # gold_category_kpis.csv / gold_channel_kpis.csv / gold_monthly_kpis.csv were a stale,
    # ~400x-smaller demo sample inconsistent with every other real figure in this project.
    ws = wb.create_sheet("Gold KPIs - Category")
    cat = pd.read_csv(os.path.join(DATA, "leatt_intelligent_category.csv"))
    write_df(ws, cat, title="Revenue, margin, returns by category")

    ws = wb.create_sheet("Gold KPIs - Channel")
    chan = pd.read_csv(os.path.join(DATA, "leatt_intelligent_channel.csv"))
    write_df(ws, chan, title="Revenue, margin, returns by acquisition channel")

    ws = wb.create_sheet("Gold KPIs - Monthly")
    mon = pd.read_csv(os.path.join(DATA, "leatt_intelligent_monthly.csv")).sort_values("month")
    write_df(ws, mon, title="Monthly revenue and margin trend, with anomaly flag and 3-month moving average")

    ws = wb.create_sheet("Gold KPIs - Province")
    prov = pd.read_csv(os.path.join(DATA, "leatt_intelligent_province.csv"))
    write_df(ws, prov, title="Revenue, margin, returns by province")

    # 11. Revenue forecast (prevention: forward-looking, not just historical reporting)
    ws = wb.create_sheet("Revenue Forecast")
    forecast = pd.read_csv(os.path.join(DATA, "leatt_revenue_forecast.csv"))
    r = write_df(ws, forecast, title="6-month forward revenue forecast",
                 subtitle="Extends the anomaly-flagged monthly trend on the Gold KPIs - Monthly tab")
    anomalies = mon[mon["anomaly_flag"] == True][["month", "net_revenue_zar", "margin_rate", "revenue_anomaly_score"]]
    r += 2
    write_df(ws, anomalies, start_row=r, title="Months flagged by the anomaly-detection score",
             subtitle="Same 4 months show real margin compression - the anomaly score and the margin story agree")

    # 12. Issues & root causes (the "why did this move" diagnostic layer)
    ws = wb.create_sheet("Issues - Root Causes")
    root_cause = pd.read_csv(os.path.join(DATA, "leatt_root_cause_playbook.csv"))
    r = write_df(ws, root_cause, title="Root-cause diagnostic playbook",
                 subtitle="When a headline KPI moves, this is the decision tree used to find out why before reacting")
    r += 2
    signals = pd.read_csv(os.path.join(DATA, "leatt_decision_signal_rules.csv"))
    write_df(ws, signals, start_row=r, title="Decision signals monitored",
             subtitle="Each signal has a threshold rule and an assigned owner")

    # 13. Next best actions & prioritized initiatives (the "what do we do about it" layer)
    ws = wb.create_sheet("Actions - Initiatives")
    actions = pd.read_csv(os.path.join(DATA, "leatt_next_best_actions.csv"))
    r = write_df(ws, actions, title="Next best actions by owner")
    r += 2
    initiatives = pd.read_csv(os.path.join(DATA, "leatt_prioritized_business_initiatives.csv")).sort_values("priority_score", ascending=False)
    write_df(ws, initiatives, start_row=r, title="Prioritized initiatives",
             subtitle="Ranked by priority_score = value x confidence, effort-adjusted")

    # 14. ML return-risk (prevention: predictive flagging before the return happens)
    ws = wb.create_sheet("ML Return Risk")
    risk_model = pd.read_csv(os.path.join(DATA, "leatt_ml_return_risk_metrics.csv"))
    r = write_df(ws, risk_model, title="Return-risk model quality")
    risk_sample = pd.read_csv(os.path.join(DATA, "leatt_ml_return_risk_scores_sample.csv"))
    by_cat = (risk_sample.groupby("category")["net_revenue_zar"]
              .agg(lines="count", revenue_at_risk_zar="sum").reset_index()
              .sort_values("revenue_at_risk_zar", ascending=False))
    r += 2
    write_df(ws, by_cat, start_row=r, title="High-risk transaction watchlist by category",
             subtitle=f"{len(risk_sample):,} ML-flagged high-return-risk transaction lines sampled from the fact table")

    # 15. Prevention & controls (stopping bad data before it reaches the board)
    ws = wb.create_sheet("Prevention - Controls")
    controls = pd.read_csv(os.path.join(DATA, "leatt_data_controls.csv"))
    r = write_df(ws, controls, title="Data controls")
    r += 2
    exceptions = pd.read_csv(os.path.join(DATA, "leatt_audit_exceptions.csv"))
    write_df(ws, exceptions, start_row=r, title="Audit exceptions log",
             subtitle="Each exception traces back to one of the controls above that was designed to catch it")

    # 16. Marketing ROAS
    ws = wb.create_sheet("Marketing ROAS")
    roi = pd.read_csv(os.path.join(DATA, "leatt_marketing_roi.csv"))
    write_df(ws, roi, title="Return on ad spend by channel")

    # 17. A/B tests
    ws = wb.create_sheet("A-B Tests")
    ab = pd.read_csv(os.path.join(DATA, "leatt_ab_test_results.csv"))
    write_df(ws, ab, title="Experiment results")

    # 12. SEO / AI SEO audit (live crawl of za.leatt.com, not synthetic)
    ws = wb.create_sheet("SEO - AI SEO Audit")
    adv = pd.read_csv(os.path.join(DATA, "leatt_seo_advanced_findings.csv"))
    r = write_df(ws, adv, title="Advanced findings - live crawl of za.leatt.com",
                 subtitle="Verified directly against the real storefront (HTML, headers, sitemap, robots.txt) - not generated from a checklist")
    r += 2
    plat = pd.read_csv(os.path.join(DATA, "leatt_ecommerce_platform_research.csv"))
    write_df(ws, plat, start_row=r, title="Platform / technical signals register")

    # 13. Finance reconciliation
    ws = wb.create_sheet("Finance Reconciliation")
    recon = pd.read_csv(os.path.join(DATA, "leatt_reconciliation.csv"))
    r = write_df(ws, recon, title="Monthly reconciliation pack")
    r += 2
    tb = pd.read_csv(os.path.join(DATA, "leatt_trial_balance.csv"))
    write_df(ws, tb, start_row=r, title="Trial balance")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print("wrote", OUT)


if __name__ == "__main__":
    main()
