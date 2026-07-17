from __future__ import annotations

import csv
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
REPO = OUT / "leatt-fabric-bi-ml-git-proof"
SQLITE = OUT / "leatt_ecommerce_warehouse.sqlite"
MASTER = OUT / "leatt_ecommerce_bi_ml_ai_command_center_master.xlsx"


def money(value: float) -> str:
    return f"R{value:,.0f}"


def pct(value: float) -> str:
    return f"{value:.1%}"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_inputs() -> dict[str, pd.DataFrame | dict[str, float | str]]:
    conn = sqlite3.connect(SQLITE)
    monthly = pd.read_sql("select * from agg_monthly order by month", conn)
    category = pd.read_sql("select * from agg_category order by net_revenue_zar desc", conn)
    channel = pd.read_sql("select * from agg_channel order by net_revenue_zar desc", conn)
    product = pd.read_sql("select * from agg_product order by net_revenue_zar desc limit 200", conn)
    total_revenue, total_margin, total_returns, orders, lines = conn.execute(
        """
        select
            sum(net_revenue_zar),
            sum(gross_margin_zar),
            sum(return_amount_zar),
            count(distinct order_id),
            count(*)
        from fact_transaction_lines
        """
    ).fetchone()
    conn.close()
    intelligent_channel = pd.read_csv(OUT / "leatt_intelligent_channel.csv")
    kpis = pd.read_csv(OUT / "leatt_kpi_rationale_catalog.csv")
    metrics = {
        "total_revenue": float(total_revenue),
        "total_margin": float(total_margin),
        "total_returns": float(total_returns),
        "orders": float(orders),
        "lines": float(lines),
        "margin_rate": float(total_margin) / float(total_revenue),
        "return_rate": float(total_returns) / float(total_revenue),
    }
    return {
        "monthly": monthly,
        "category": category,
        "channel": channel,
        "intelligent_channel": intelligent_channel,
        "product": product,
        "kpis": kpis,
        "metrics": metrics,
    }


def decision_signal_rows(data: dict[str, pd.DataFrame | dict[str, float | str]]) -> pd.DataFrame:
    metrics = data["metrics"]
    category = data["category"].copy()
    ichannel = data["intelligent_channel"].copy()
    monthly = data["monthly"].copy()
    category["return_rate"] = category["return_amount_zar"] / category["net_revenue_zar"]
    category["margin_rate"] = category["gross_margin_zar"] / category["net_revenue_zar"]
    top_category = category.iloc[0]
    risky_category = category.sort_values("return_rate", ascending=False).iloc[0]
    best_channel = ichannel.sort_values("roas", ascending=False).iloc[0]
    weak_channel = ichannel.sort_values("roas", ascending=True).iloc[0]
    monthly["revenue_ma3"] = monthly["net_revenue_zar"].rolling(3, min_periods=2).mean()
    monthly["movement"] = ((monthly["net_revenue_zar"] - monthly["revenue_ma3"]) / monthly["revenue_ma3"]).fillna(0)
    last_anomaly = monthly.loc[monthly["movement"].abs() > 0.12].tail(1)
    anomaly_signal = "No current material anomaly"
    if not last_anomaly.empty:
        anomaly = last_anomaly.iloc[0]
        anomaly_signal = f"{anomaly['month']} moved {anomaly['movement']:.1%} versus 3-month signal"

    rows = [
        ("Revenue momentum", money(metrics["total_revenue"]), "Healthy scale", "If monthly movement exceeds +/-12%, explain by category, channel and stock.", anomaly_signal, "BI Ops", "Trigger exception commentary in the monthly trading pack."),
        ("Margin quality", pct(metrics["margin_rate"]), "Strong but must be protected", "If margin rate falls below 45%, inspect discounting, COGS and channel mix.", "Growth is valuable because modeled margin is still high.", "CFO / Merchandising", "Protect margin-positive product mix and bundle high-margin items."),
        ("Return leakage", pct(metrics["return_rate"]), "Watchlist", "If return rate exceeds 5%, prioritize category root-cause work before scaling media.", f"{risky_category['category']} has the highest return-rate signal.", "COO / CX", "Improve sizing, delivery promise, PDP content and post-purchase guidance."),
        ("Hero category concentration", f"{top_category['category']} {money(top_category['net_revenue_zar'])}", "Strategic dependency", "If one category exceeds 35% share, manage stockout and campaign dependency risk.", "Apparel is a growth engine but also a concentration risk.", "CEO / Supply", "Create availability alerts and category-specific SEO/story pages."),
        ("Channel efficiency", f"{best_channel['channel']} {best_channel['roas']:.1f}x ROAS", "Scale carefully", "Scale only when ROAS, margin and conversion all stay above threshold.", f"{weak_channel['channel']} is the lowest-ROAS channel and should stay experimental.", "CMO", "Move spend toward efficient journeys; keep paid experiments controlled."),
        ("Data completeness", f"{metrics['lines']:,.0f} rows", "Audit-ready", "If row counts differ from expected load by >1%, block semantic model promotion.", "Million-row scale is loaded and traceable to OneLake/Git proof.", "Data Governance", "Publish source, target, row count and commit ID in every close pack."),
        ("Cloud cost", "Fabric F2 paused", "Controlled", "If no active demo or refresh, capacity must be paused.", "Capacity was paused after handover to protect remaining credit.", "BI Platform Owner", "Resume only for refresh/demo, then pause again."),
    ]
    return pd.DataFrame(rows, columns=["signal", "current_reading", "health", "threshold_or_rule", "what_it_means", "owner", "recommended_decision"])


def initiative_rows(data: dict[str, pd.DataFrame | dict[str, float | str]]) -> pd.DataFrame:
    metrics = data["metrics"]
    category = data["category"].copy()
    product = data["product"].copy()
    ichannel = data["intelligent_channel"].copy()
    category["return_rate"] = category["return_amount_zar"] / category["net_revenue_zar"]
    product["margin_rate"] = product["gross_margin_zar"] / product["net_revenue_zar"]
    risky_category = category.sort_values("return_rate", ascending=False).iloc[0]
    hero_category = category.iloc[0]
    margin_product = product.sort_values("gross_margin_zar", ascending=False).iloc[0]
    best_channel = ichannel.sort_values("roas", ascending=False).iloc[0]

    initiatives = [
        {
            "initiative": "Return leakage reduction sprint",
            "hypothesis": f"{risky_category['category']} returns can be reduced with fit guidance, PDP content and post-purchase education.",
            "evidence": f"Portfolio return leakage is {pct(metrics['return_rate'])}; {risky_category['category']} is the highest-risk category.",
            "estimated_value": money(metrics["total_returns"] * 0.10),
            "confidence": "Medium",
            "effort": "Medium",
            "priority_score": 92,
            "owner": "COO / CX",
            "next_step": "Create return-reason taxonomy and product-level return dashboard.",
        },
        {
            "initiative": "Hero category availability and SEO program",
            "hypothesis": f"{hero_category['category']} deserves protected stock, SEO hubs and executive tracking.",
            "evidence": f"{hero_category['category']} contributes {money(hero_category['net_revenue_zar'])} modeled revenue.",
            "estimated_value": money(hero_category["net_revenue_zar"] * 0.015 * metrics["margin_rate"]),
            "confidence": "High",
            "effort": "Medium",
            "priority_score": 90,
            "owner": "CEO / Ecommerce",
            "next_step": "Build category scorecard with stockout, ranking, margin and landing page signals.",
        },
        {
            "initiative": "Margin-positive bundle testing",
            "hypothesis": f"Bundling {margin_product['product_title'][:70]} with accessories can increase AOV without margin dilution.",
            "evidence": "AOV and gross margin are the cleanest basket-quality measures.",
            "estimated_value": money(metrics["total_margin"] * 0.01),
            "confidence": "Medium",
            "effort": "Low",
            "priority_score": 88,
            "owner": "Merchandising / CMO",
            "next_step": "Run A/B test: product page bundle versus normal recommendation carousel.",
        },
        {
            "initiative": "Channel budget guardrails",
            "hypothesis": f"{best_channel['channel']} is efficient, but paid budget should expand only through controlled tests.",
            "evidence": f"Best observed ROAS signal is {best_channel['roas']:.1f}x.",
            "estimated_value": money(metrics["total_revenue"] * 0.004 * metrics["margin_rate"]),
            "confidence": "Medium",
            "effort": "Low",
            "priority_score": 84,
            "owner": "CMO",
            "next_step": "Set campaign guardrails: minimum ROAS, conversion lift and margin contribution.",
        },
        {
            "initiative": "SAP close reconciliation automation",
            "hypothesis": "Order/refund/VAT/payment exceptions can be reduced by automated SAP B1/BW reconciliation.",
            "evidence": "Finance pack already maps ecommerce facts to accounting controls and exception registers.",
            "estimated_value": "Risk reduction / faster close",
            "confidence": "High",
            "effort": "Medium",
            "priority_score": 82,
            "owner": "CFO / BI Lead",
            "next_step": "Schedule monthly exception pack with owner, SLA and SAP account mapping.",
        },
    ]
    return pd.DataFrame(initiatives)


def root_cause_rows() -> pd.DataFrame:
    rows = [
        ("Revenue down", "Category demand", "Which categories moved versus 3-month average?", "MonthlyTrend + CategoryKPI", "Campaign issue, product availability, seasonality or assortment gap."),
        ("Revenue down", "Channel mix", "Did traffic/revenue shift away from efficient channels?", "ChannelROI", "Media allocation, SEO ranking or landing-page relevance issue."),
        ("Margin down", "Discounting", "Did discount depth increase by category/channel?", "Fact transactions + ProductKPI", "Promo overuse or weak pricing control."),
        ("Margin down", "Returns", "Are returns rising in a specific category/product?", "MLReturnRisk + Accounting Pack", "Sizing, content, fulfillment or quality leakage."),
        ("Returns up", "Product fit", "Are high-risk items clustered by category, size or product line?", "ML scores + ProductKPI", "Improve size guide, PDP detail and customer education."),
        ("Returns up", "Fulfillment", "Are return rates linked to fulfillment days or province?", "Fact transactions + ProvinceKPI", "Delivery promise or regional service gap."),
        ("ROAS down", "Traffic quality", "Is revenue falling while spend remains high?", "MarketingROI", "Creative fatigue, wrong audience, poor landing page."),
        ("Audit exception", "Data lineage", "Do source/target row counts reconcile?", "Source Register + Fabric Proof", "Duplicate, missing or stale pipeline load."),
    ]
    return pd.DataFrame(rows, columns=["problem", "diagnostic_branch", "question_to_ask", "data_needed", "likely_interpretation"])


def executive_story_rows(signals: pd.DataFrame, initiatives: pd.DataFrame) -> pd.DataFrame:
    rows = [
        ("Board message", "The project has moved beyond reporting into an operating model: every KPI has a signal rule, owner and decision.", "Use this to demonstrate senior BI leadership, not dashboard production."),
        ("Commercial message", "Growth is meaningful only if margin, return leakage and channel economics stay healthy.", "Open with revenue, then immediately explain margin and returns."),
        ("ML message", "The return-risk model is a recall-focused watchlist, not a magic answer.", "This is mature: it says how the model should and should not be used."),
        ("Governance message", "The row counts, OneLake paths, Git commits and paused capacity make the work auditable.", "This answers 'prove it' and 'what did it cost?' in the same story."),
        ("Next decision", initiatives.iloc[0]["initiative"], initiatives.iloc[0]["next_step"]),
    ]
    return pd.DataFrame(rows, columns=["story_moment", "intelligent_position", "how_to_say_it"])


def markdown_table(df: pd.DataFrame) -> str:
    rows = ["| " + " | ".join(df.columns) + " |", "| " + " | ".join(["---"] * len(df.columns)) + " |"]
    for _, row in df.iterrows():
        rows.append("| " + " | ".join(str(row[col]).replace("\n", " ") for col in df.columns) + " |")
    return "\n".join(rows)


def write_outputs(signals: pd.DataFrame, initiatives: pd.DataFrame, root_causes: pd.DataFrame, story: pd.DataFrame) -> list[Path]:
    outputs = []
    for name, df in [
        ("leatt_decision_signal_rules.csv", signals),
        ("leatt_prioritized_business_initiatives.csv", initiatives),
        ("leatt_root_cause_playbook.csv", root_causes),
        ("leatt_executive_storyline.csv", story),
    ]:
        path = OUT / name
        df.to_csv(path, index=False)
        outputs.append(path)

    md = OUT / "decision_intelligence_playbook.md"
    md.write_text(
        f"""# Decision Intelligence Playbook

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}

This layer turns the BI project into a management system. It defines what each signal means, when to act, who owns the action, and how to explain the decision.

## Signal Rules

{markdown_table(signals)}

## Prioritized Business Initiatives

{markdown_table(initiatives)}

## Root-Cause Playbook

{markdown_table(root_causes)}

## Executive Storyline

{markdown_table(story)}
""",
        encoding="utf-8",
    )
    outputs.append(md)

    pdf = OUT / "leatt_decision_intelligence_playbook.pdf"
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleRed", parent=styles["Title"], alignment=TA_CENTER, textColor=colors.HexColor("#d71920")))
    pdf_story = [
        Paragraph("Decision Intelligence Playbook", styles["TitleRed"]),
        Paragraph("Signal rules, root-cause logic, owner actions and prioritized business initiatives.", styles["BodyText"]),
        Spacer(1, 0.35 * cm),
        Table(
            [list(signals.columns)] + signals.values.tolist(),
            colWidths=[3.2 * cm, 3.0 * cm, 2.6 * cm, 5.0 * cm, 5.0 * cm, 3.0 * cm, 5.0 * cm],
            repeatRows=1,
            style=[
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
                ("FONTSIZE", (0, 0), (-1, -1), 6.3),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ],
        ),
        PageBreak(),
        Paragraph("Prioritized Initiatives", styles["Heading1"]),
        Table(
            [list(initiatives.columns)] + initiatives.values.tolist(),
            colWidths=[3.8 * cm, 5.0 * cm, 5.0 * cm, 2.5 * cm, 2.2 * cm, 2.0 * cm, 2.2 * cm, 3.0 * cm, 4.0 * cm],
            repeatRows=1,
            style=[
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
                ("FONTSIZE", (0, 0), (-1, -1), 5.5),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ],
        ),
        PageBreak(),
        Paragraph("Root-Cause Playbook", styles["Heading1"]),
        Table(
            [list(root_causes.columns)] + root_causes.values.tolist(),
            colWidths=[3.3 * cm, 3.5 * cm, 6.0 * cm, 5.0 * cm, 7.0 * cm],
            repeatRows=1,
            style=[
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#dddddd")),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ],
        ),
    ]
    SimpleDocTemplate(str(pdf), pagesize=landscape(A4), rightMargin=0.7 * cm, leftMargin=0.7 * cm, topMargin=0.7 * cm, bottomMargin=0.7 * cm).build(pdf_story)
    outputs.append(pdf)
    return outputs


def update_workbook(signals: pd.DataFrame, initiatives: pd.DataFrame, root_causes: pd.DataFrame, story: pd.DataFrame) -> Path:
    wb = load_workbook(MASTER)
    for sheet in ["Decision Signals", "Business Initiatives", "Root Cause Playbook", "Executive Storyline"]:
        if sheet in wb.sheetnames:
            del wb[sheet]
    for sheet, df in [
        ("Decision Signals", signals),
        ("Business Initiatives", initiatives),
        ("Root Cause Playbook", root_causes),
        ("Executive Storyline", story),
    ]:
        ws = wb.create_sheet(sheet)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="111111")
        ws.freeze_panes = "A2"
        for col in "ABCDEFGHIJKL":
            ws.column_dimensions[col].width = 32
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(MASTER)
    return MASTER


def copy_to_repo(paths: list[Path], master: Path) -> None:
    for folder in [REPO / "docs", REPO / "artifacts" / "data_samples", REPO / "artifacts" / "reports"]:
        folder.mkdir(parents=True, exist_ok=True)
    for path in paths:
        if path.suffix.lower() == ".md":
            shutil.copy2(path, REPO / "docs" / path.name)
        elif path.suffix.lower() == ".csv":
            shutil.copy2(path, REPO / "artifacts" / "data_samples" / path.name)
        elif path.suffix.lower() == ".pdf":
            shutil.copy2(path, REPO / "artifacts" / "reports" / path.name)
    shutil.copy2(master, REPO / "artifacts" / "reports" / master.name)


def main() -> None:
    data = load_inputs()
    signals = decision_signal_rows(data)
    initiatives = initiative_rows(data)
    root_causes = root_cause_rows()
    story = executive_story_rows(signals, initiatives)
    paths = write_outputs(signals, initiatives, root_causes, story)
    master = update_workbook(signals, initiatives, root_causes, story)
    copy_to_repo(paths, master)
    print("Decision intelligence layer complete:")
    for path in [*paths, master]:
        print(f"- {path} ({path.stat().st_size / (1024 * 1024):.2f} MB)")


if __name__ == "__main__":
    main()
