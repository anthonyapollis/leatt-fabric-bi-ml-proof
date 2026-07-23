from __future__ import annotations

import csv
import html
import json
import shutil
import sqlite3
import zipfile
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Image as PdfImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table


ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"
REPO = OUT / "leatt-fabric-bi-ml-git-proof"
SCREENSHOTS = OUT / "screenshots"
SQLITE = OUT / "leatt_ecommerce_warehouse.sqlite"


def money(value: float) -> str:
    return f"R{value:,.0f}"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def load_metrics() -> dict[str, str]:
    story = read_csv(OUT / "leatt_data_story_business_optimization.csv")
    ml = read_csv(OUT / "leatt_ml_return_risk_metrics.csv")[0]
    proof = json.loads((OUT / "powerbi_semantic_model_proof.json").read_text(encoding="utf-8"))
    conn = sqlite3.connect(SQLITE)
    row_count = conn.execute("select count(*) from fact_transaction_lines").fetchone()[0]
    product_count = conn.execute("select count(*) from dim_product_variant").fetchone()[0]
    customer_count = conn.execute("select count(*) from dim_customer").fetchone()[0]
    total_revenue, total_margin, total_returns = conn.execute(
        "select sum(net_revenue_zar), sum(gross_margin_zar), sum(return_amount_zar) from fact_transaction_lines"
    ).fetchone()
    conn.close()
    dataset_url = proof.get(
        "dataset_url",
        f"https://app.powerbi.com/groups/{proof['workspace_id']}/datasets/{proof['dataset_id']}",
    )
    return {
        "row_count": f"{row_count:,}",
        "product_count": f"{product_count:,}",
        "customer_count": f"{customer_count:,}",
        "total_revenue": money(total_revenue),
        "total_margin": money(total_margin),
        "return_rate": f"{total_returns / total_revenue:.1%}",
        "ml_auc": f"{float(ml['auc']):.3f}",
        "ml_recall": f"{float(ml['recall']):.3f}",
        "dataset_id": proof["dataset_id"],
        "dataset_url": dataset_url,
        "story": story,
    }


def write_markdown(metrics: dict[str, str]) -> Path:
    story_rows = "\n".join(
        f"- **{row['theme']}**: {row['what_the_data_tells_us']}" for row in metrics["story"]
    )
    path = OUT / "final_project_completion_report.md"
    path.write_text(
        f"""# Leatt Fabric BI/ML Project Completion Report

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}

## Executive conclusion

This project is now a complete senior BI portfolio case study: website extraction, million-row ecommerce modeling, Azure/Fabric Lakehouse upload, Data Factory proof, Power BI semantic model, ML return-risk scoring, accounting reconciliation, SAP integration design, governance controls, A/B testing, marketing/SEO analysis, AI agents, Git proof and executive reporting.

## What was built

- Extracted real Leatt ZA public catalog/product evidence and confirmed the storefront platform pattern.
- Generated and modeled {metrics['row_count']} ecommerce transaction lines, {metrics['product_count']} product records and {metrics['customer_count']} synthetic customers.
- Uploaded the data estate to Microsoft Fabric OneLake under the `Apollis` workspace and `Leatt_BI_ML_Lakehouse`.
- Created a Fabric Data Factory pipeline item and Power BI semantic model proof.
- Built Excel/PDF/eBook deliverables, evidence screenshots, ERD, data governance register, accounting reconciliation pack and AI command-center layer.
- Pushed the Git delivery package to GitHub: `https://github.com/anthonyapollis/leatt-fabric-bi-ml-proof`.

## What the data tells us

{story_rows}

## Business optimisation thesis

- Protect hero categories with stock, price, SEO and executive alerts.
- Use A/B tests to prove bundles, free shipping thresholds and lifecycle offers before scaling spend.
- Treat returns as a margin problem, not only an operations problem: improve fit guidance, product detail pages, delivery promises and post-purchase education.
- Reconcile ecommerce orders, refunds, VAT, payment settlement and SAP postings in the monthly close pack.
- Promote Fabric/Power BI as the trusted semantic layer and keep source lineage, row counts, file hashes and Git commits visible.

## Azure/Fabric proof

- Resource group: `rg-leatt-fabric-bi-ml`
- Fabric capacity: `fabric-capacity-redacted`
- Capacity SKU: `F2`
- Workspace: `Apollis`
- Lakehouse: `Leatt_BI_ML_Lakehouse`
- Pipeline: `pipeline name redacted for public sharing`
- Power BI semantic model ID: `{metrics['dataset_id']}`
- Power BI semantic model URL: `{metrics['dataset_url']}`

## ML proof

- Return-risk model AUC: `{metrics['ml_auc']}`
- Return-risk model recall: `{metrics['ml_recall']}`
- Interpretation: the model is useful as an early recall-focused watchlist, but precision should be improved with real return reasons, size/fit attributes, delivery events and customer service signals.

## SAP and platform positioning

- The live storefront evidence points to Shopify, so SAP Commerce Cloud / Hybris is not required for this specific Leatt ZA storefront proof.
- SAP Business One or SAP BW still fits as the downstream source for finance, inventory, VAT, COGS, customer/accounting reconciliation and audit extracts.
- SAP Commerce Cloud / Hybris would fit if the company wanted a larger enterprise commerce platform with deep SAP ERP integration, complex catalogs, B2B pricing and global multi-site orchestration.
- dbt is optional for this proof. It becomes valuable later for versioned SQL transformations, automated tests and CI/CD around the semantic warehouse layer.

## Cost and handover

Suspend Fabric capacity when review is complete:

```powershell
az fabric capacity suspend --resource-group <resource-group> --capacity-name <capacity-name>
```

Resume only when refreshing Fabric/Power BI proof:

```powershell
az fabric capacity resume --resource-group <resource-group> --capacity-name <capacity-name>
```
""",
        encoding="utf-8",
    )
    return path


def write_pdf(metrics: dict[str, str]) -> Path:
    path = OUT / "leatt_final_executive_portfolio_dossier.pdf"
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="TitleRed", parent=styles["Title"], alignment=TA_CENTER, textColor=colors.HexColor("#d71920")))
    styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=8, leading=10))
    story = [
        Paragraph("Leatt Fabric BI/ML Executive Portfolio Dossier", styles["TitleRed"]),
        Paragraph("End-to-end ecommerce BI, ML, accounting governance, Fabric, Power BI and AI command-center proof.", styles["BodyText"]),
        Spacer(1, 0.4 * cm),
        Table(
            [
                ["Metric", "Value"],
                ["Transaction lines", metrics["row_count"]],
                ["Products", metrics["product_count"]],
                ["Customers", metrics["customer_count"]],
                ["Revenue", metrics["total_revenue"]],
                ["Gross margin", metrics["total_margin"]],
                ["Return leakage", metrics["return_rate"]],
                ["ML AUC / Recall", f"{metrics['ml_auc']} / {metrics['ml_recall']}"],
            ],
            colWidths=[7 * cm, 9 * cm],
            style=[
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ],
        ),
    ]
    story.append(PageBreak())
    for title, image_name in [
        ("Power BI Executive Overview", "powerbi_executive_overview.png"),
        ("Power BI ML Monitoring", "powerbi_ml_monitoring.png"),
        ("Marketing and Fabric Proof", "powerbi_marketing_fabric_proof.png"),
        ("AI Commerce Command Center", "ai_commerce_command_center.png"),
        ("Merchandising and Governance", "intelligent_merchandising_governance.png"),
    ]:
        img = SCREENSHOTS / image_name
        if img.exists():
            story.extend([Paragraph(title, styles["Heading1"]), PdfImage(str(img), width=24.5 * cm, height=13.8 * cm), PageBreak()])
    story.extend(
        [
            Paragraph("What The Data Tells Us", styles["Heading1"]),
            Table(
                [["Theme", "Insight"]]
                + [[row["theme"], row["what_the_data_tells_us"]] for row in metrics["story"]],
                colWidths=[4.2 * cm, 20 * cm],
                repeatRows=1,
                style=[
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111111")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ],
            ),
            Spacer(1, 0.4 * cm),
            Paragraph("Handover", styles["Heading1"]),
            Paragraph("GitHub proof: https://github.com/anthonyapollis/leatt-fabric-bi-ml-proof", styles["BodyText"]),
            Paragraph("Suspend Fabric Fabric capacity after review to protect remaining Azure credit.", styles["BodyText"]),
        ]
    )
    SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=1 * cm, leftMargin=1 * cm, topMargin=1 * cm, bottomMargin=1 * cm).build(story)
    return path


def write_artifact_manifest(metrics: dict[str, str]) -> Path:
    path = OUT / "leatt_final_artifact_manifest.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Portfolio Index"
    ws.append(["artifact", "path", "purpose"])
    artifacts = [
        ("Final dossier", OUT / "leatt_final_executive_portfolio_dossier.pdf", "Executive project summary"),
        ("AI command center report", OUT / "ai_commerce_command_center_report.pdf", "AI operating model and next-best actions"),
        ("Master Excel", OUT / "leatt_ecommerce_bi_ml_ai_command_center_master.xlsx", "Workbook with BI, ERD, Fabric, Power BI and data story sheets"),
        ("Power BI screenshots", OUT / "leatt_powerbi_ml_report_screenshots.pdf", "Report screenshots/evidence pack"),
        ("Accounting governance pack", OUT / "leatt_accounting_reconciliation_governance_sap_pack.xlsx", "SAP/accounting reconciliation and controls"),
        ("GitHub repo", Path("https://github.com/anthonyapollis/leatt-fabric-bi-ml-proof"), "Remote proof repository"),
    ]
    for name, artifact_path, purpose in artifacts:
        ws.append([name, str(artifact_path), purpose])
    ws2 = wb.create_sheet("Key Metrics")
    for key, value in metrics.items():
        if key != "story":
            ws2.append([key, value])
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="111111")
        sheet.freeze_panes = "A2"
        for col in "ABCDEF":
            sheet.column_dimensions[col].width = 42
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)
    return path


def ppt_text_box(x: int, y: int, w: int, h: int, text: str, font_size: int = 2800, bold: bool = False) -> str:
    b = "<a:b/>" if bold else ""
    lines = text.split("\n")
    paras = "".join(
        f"<a:p><a:r><a:rPr lang=\"en-US\" sz=\"{font_size}\">{b}<a:solidFill><a:srgbClr val=\"111111\"/></a:solidFill></a:rPr><a:t>{escape(line)}</a:t></a:r></a:p>"
        for line in lines
    )
    return f"""
    <p:sp><p:nvSpPr><p:cNvPr id=\"2\" name=\"TextBox\"/><p:cNvSpPr txBox=\"1\"/><p:nvPr/></p:nvSpPr>
    <p:spPr><a:xfrm><a:off x=\"{x}\" y=\"{y}\"/><a:ext cx=\"{w}\" cy=\"{h}\"/></a:xfrm><a:prstGeom prst=\"rect\"><a:avLst/></a:prstGeom><a:noFill/><a:ln><a:noFill/></a:ln></p:spPr>
    <p:txBody><a:bodyPr wrap=\"square\"/><a:lstStyle/>{paras}</p:txBody></p:sp>"""


def ppt_image(x: int, y: int, w: int, h: int, rid: str) -> str:
    return f"""
    <p:pic><p:nvPicPr><p:cNvPr id=\"5\" name=\"Picture\"/><p:cNvPicPr/><p:nvPr/></p:nvPicPr>
    <p:blipFill><a:blip r:embed=\"{rid}\"/><a:stretch><a:fillRect/></a:stretch></p:blipFill>
    <p:spPr><a:xfrm><a:off x=\"{x}\" y=\"{y}\"/><a:ext cx=\"{w}\" cy=\"{h}\"/></a:xfrm><a:prstGeom prst=\"rect\"><a:avLst/></a:prstGeom></p:spPr></p:pic>"""


def slide_xml(title: str, body: str, image_rid: str | None = None) -> str:
    image = ppt_image(457200, 1752600, 8229600, 4629150, image_rid) if image_rid else ""
    body_box = ppt_text_box(685800, 1752600, 7772400, 3657600, body, 2200) if not image_rid else ""
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">
<p:cSld><p:spTree><p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr><p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/><a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr>
{ppt_text_box(457200, 342900, 8229600, 914400, title, 3600, True)}
{body_box}{image}
</p:spTree></p:cSld><p:clrMapOvr><a:masterClrMapping/></p:clrMapOvr></p:sld>"""


def make_pptx(metrics: dict[str, str]) -> Path:
    path = OUT / "leatt_fabric_bi_ml_board_presentation.pptx"
    slides = [
        ("Leatt Fabric BI/ML Portfolio", f"{metrics['row_count']} transaction lines\n{metrics['total_revenue']} revenue\nFabric + Power BI + ML + SAP governance + AI agents", None),
        ("What The Data Tells Us", "\n".join(f"{r['theme']}: {r['what_the_data_tells_us']}" for r in metrics["story"][:4]), None),
        ("Power BI Executive Overview", "", "powerbi_executive_overview.png"),
        ("ML Return Risk Monitoring", "", "powerbi_ml_monitoring.png"),
        ("AI Commerce Command Center", "", "ai_commerce_command_center.png"),
        ("Final Handover", "GitHub: https://github.com/anthonyapollis/leatt-fabric-bi-ml-proof\nSuspend Fabric Fabric capacity after review.\nUse the master Excel and dossier as project handover artifacts.", None),
    ]
    image_files = [s[2] for s in slides if s[2]]
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        z.writestr("[Content_Types].xml", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Default Extension="png" ContentType="image/png"/><Override PartName="/ppt/presentation.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.presentation.main+xml"/><Override PartName="/ppt/slideMasters/slideMaster1.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slideMaster+xml"/><Override PartName="/ppt/slideLayouts/slideLayout1.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slideLayout+xml"/>""" + "".join(f'<Override PartName="/ppt/slides/slide{i}.xml" ContentType="application/vnd.openxmlformats-officedocument.presentationml.slide+xml"/>' for i in range(1, len(slides) + 1)) + "</Types>")
        z.writestr("_rels/.rels", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="ppt/presentation.xml"/></Relationships>""")
        slide_ids = "".join(f'<p:sldId id="{255+i}" r:id="rId{i}"/>' for i in range(1, len(slides) + 1))
        z.writestr("ppt/presentation.xml", f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?><p:presentation xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"><p:sldMasterIdLst><p:sldMasterId id="2147483648" r:id="rId{len(slides)+1}"/></p:sldMasterIdLst><p:sldIdLst>{slide_ids}</p:sldIdLst><p:sldSz cx="9144000" cy="5143500" type="screen16x9"/><p:notesSz cx="6858000" cy="9144000"/></p:presentation>""")
        rels = "".join(f'<Relationship Id="rId{i}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slide" Target="slides/slide{i}.xml"/>' for i in range(1, len(slides) + 1))
        rels += f'<Relationship Id="rId{len(slides)+1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slideMaster" Target="slideMasters/slideMaster1.xml"/>'
        z.writestr("ppt/_rels/presentation.xml.rels", f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{rels}</Relationships>""")
        z.writestr("ppt/slideMasters/slideMaster1.xml", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?><p:sldMaster xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"><p:cSld><p:spTree><p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr><p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/><a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr></p:spTree></p:cSld><p:clrMap bg1="lt1" tx1="dk1" bg2="lt2" tx2="dk2" accent1="accent1" accent2="accent2" accent3="accent3" accent4="accent4" accent5="accent5" accent6="accent6" hlink="hlink" folHlink="folHlink"/><p:sldLayoutIdLst><p:sldLayoutId id="2147483649" r:id="rId1"/></p:sldLayoutIdLst></p:sldMaster>""")
        z.writestr("ppt/slideMasters/_rels/slideMaster1.xml.rels", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slideLayout" Target="../slideLayouts/slideLayout1.xml"/></Relationships>""")
        z.writestr("ppt/slideLayouts/slideLayout1.xml", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?><p:sldLayout xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main" type="blank"><p:cSld name="Blank"><p:spTree><p:nvGrpSpPr><p:cNvPr id="1" name=""/><p:cNvGrpSpPr/><p:nvPr/></p:nvGrpSpPr><p:grpSpPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="0" cy="0"/><a:chOff x="0" y="0"/><a:chExt cx="0" cy="0"/></a:xfrm></p:grpSpPr></p:spTree></p:cSld><p:clrMapOvr><a:masterClrMapping/></p:clrMapOvr></p:sldLayout>""")
        z.writestr("ppt/slideLayouts/_rels/slideLayout1.xml.rels", """<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slideMaster" Target="../slideMasters/slideMaster1.xml"/></Relationships>""")
        image_index = 1
        for i, (title, body, image_name) in enumerate(slides, start=1):
            rid = None
            if image_name:
                rid = "rId2"
                z.write(SCREENSHOTS / image_name, f"ppt/media/image{image_index}.png")
            z.writestr(f"ppt/slides/slide{i}.xml", slide_xml(title, body, rid))
            rel = '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/slideLayout" Target="../slideLayouts/slideLayout1.xml"/>'
            if image_name:
                rel += f'<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/image{image_index}.png"/>'
                image_index += 1
            z.writestr(f"ppt/slides/_rels/slide{i}.xml.rels", f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{rel}</Relationships>""")
    return path


def write_html_index(metrics: dict[str, str]) -> Path:
    path = OUT / "portfolio_index.html"
    repo = "leatt-fabric-bi-ml-git-proof"
    links = [
        ("Final Executive Dossier", "leatt_final_executive_portfolio_dossier.pdf"),
        ("Board Presentation", "leatt_fabric_bi_ml_board_presentation.pptx"),
        ("Master Excel Workbook", "leatt_ecommerce_bi_ml_ai_command_center_master.xlsx"),
        ("AI Command Center Report", "ai_commerce_command_center_report.pdf"),
        ("Power BI / ML Screenshots PDF", "leatt_powerbi_ml_report_screenshots.pdf"),
        ("Git Proof Repo Folder", repo + "/README.md"),
    ]
    cards = "\n".join(f'<a class="card" href="{html.escape(href)}"><strong>{html.escape(label)}</strong><span>{html.escape(href)}</span></a>' for label, href in links)
    story = "\n".join(f"<li><strong>{html.escape(r['theme'])}</strong>: {html.escape(r['what_the_data_tells_us'])}</li>" for r in metrics["story"])
    path.write_text(
        f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>Leatt Fabric BI/ML Portfolio Index</title>
  <style>
    body {{ font-family: Arial, sans-serif; margin: 0; background: #f6f7f9; color: #111; }}
    header {{ background: #111; color: white; padding: 36px 48px; border-bottom: 5px solid #d71920; }}
    main {{ max-width: 1120px; margin: 0 auto; padding: 32px; }}
    .grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 14px; }}
    .card {{ display: block; padding: 18px; border: 1px solid #ddd; background: white; text-decoration: none; color: #111; border-radius: 6px; }}
    .card span {{ display: block; margin-top: 8px; color: #555; font-size: 12px; word-break: break-word; }}
    li {{ margin: 10px 0; }}
  </style>
</head>
<body>
  <header>
    <h1>Leatt Fabric BI/ML Portfolio</h1>
    <p>{metrics['row_count']} transaction lines | {metrics['total_revenue']} revenue | Fabric + Power BI + ML + SAP governance + AI agents</p>
  </header>
  <main>
    <h2>Open the deliverables</h2>
    <section class="grid">{cards}</section>
    <h2>What the data tells us</h2>
    <ul>{story}</ul>
    <h2>Cost control</h2>
    <p>After review, suspend Fabric capacity: <code>az fabric capacity suspend --resource-group <resource-group> --capacity-name <capacity-name></code></p>
  </main>
</body>
</html>""",
        encoding="utf-8",
    )
    return path


def write_handover_docs(metrics: dict[str, str]) -> list[Path]:
    checklist = OUT / "final_project_completion_checklist.md"
    checklist.write_text(
        """# Final Project Completion Checklist

- [x] Real ecommerce catalog/product source captured and documented.
- [x] Ecommerce platform research completed: Shopify signals, SAP Commerce/Hybris positioning explained.
- [x] Millions-row transaction dataset generated and modeled.
- [x] Fabric capacity/workspace/lakehouse/pipeline proof captured.
- [x] OneLake upload proof created for transactions, catalog, customers and ML scores.
- [x] Power BI semantic model created with executive, marketing, ML and proof tables.
- [x] ML return-risk model trained and documented.
- [x] A/B testing, marketing ROI, competitor and SEO analysis included.
- [x] Accounting reconciliation, SAP B1/BW mapping, VAT and governance controls included.
- [x] ERD/star schema included.
- [x] Excel, PDF, eBook, screenshot pack, presentation and HTML portfolio index generated.
- [x] GitHub private proof repository created and pushed.
- [ ] Optional next step: create a native Power BI report in the service from the semantic model if UI/browser access is stable.
- [ ] Optional next step: convert uploaded OneLake files to managed Delta tables and schedule production refresh.
""",
        encoding="utf-8",
    )
    cost = OUT / "fabric_cost_control_and_handover.md"
    cost.write_text(
        """# Fabric Cost Control And Handover

The Fabric Fabric capacity is the main ongoing cost item. Keep it running only while proving or refreshing the project.

Suspend:

```powershell
az fabric capacity suspend --resource-group <resource-group> --capacity-name <capacity-name>
```

Resume:

```powershell
az fabric capacity resume --resource-group <resource-group> --capacity-name <capacity-name>
```

Recommended next production hardening:

- Convert Lakehouse `Files/Bronze` parquet/CSV files to Delta tables.
- Add Data Factory pipeline activity logging and scheduled refresh.
- Add row-count, null-check and reconciliation tests.
- Restrict access with workspace roles and PII minimization.
- Create a monthly close evidence pack with source paths, row counts, Git commit and semantic model refresh time.
""",
        encoding="utf-8",
    )
    return [checklist, cost]


def copy_to_repo(paths: list[Path]) -> None:
    (REPO / "docs").mkdir(parents=True, exist_ok=True)
    (REPO / "artifacts" / "reports").mkdir(parents=True, exist_ok=True)
    (REPO / "artifacts" / "data_samples").mkdir(parents=True, exist_ok=True)
    for path in paths:
        if path.suffix.lower() in {".md", ".html"}:
            shutil.copy2(path, REPO / "docs" / path.name)
        elif path.suffix.lower() in {".pdf", ".xlsx", ".pptx"}:
            shutil.copy2(path, REPO / "artifacts" / "reports" / path.name)
        elif path.suffix.lower() in {".csv", ".json"}:
            shutil.copy2(path, REPO / "artifacts" / "data_samples" / path.name)


def refresh_zip() -> Path:
    zip_base = OUT / "leatt-fabric-bi-ml-git-proof"
    zip_path = OUT / "leatt-fabric-bi-ml-git-proof.zip"
    if zip_path.exists():
        zip_path.unlink()
    shutil.make_archive(str(zip_base), "zip", REPO)
    return zip_path


def main() -> None:
    metrics = load_metrics()
    paths = [
        write_markdown(metrics),
        write_pdf(metrics),
        write_artifact_manifest(metrics),
        make_pptx(metrics),
        write_html_index(metrics),
        *write_handover_docs(metrics),
    ]
    copy_to_repo(paths)
    zip_path = refresh_zip()
    print("Final project pack complete:")
    for path in [*paths, zip_path]:
        print(f"- {path} ({path.stat().st_size / (1024 * 1024):.2f} MB)")


if __name__ == "__main__":
    main()
